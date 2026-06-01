"""
Dashboard Boilerplate — EKOMAT Analytics Hub.
FastAPI + asyncpg + Jinja2 + WooCommerce + PrestaShop + Allegro.

Сайты:
  - dobraszklarnia.pl  (WooCommerce)  — webhook + API sync
  - oteko.pl           (PrestaShop)   — API pull каждые 15 мин
  - cieplarnia.pl      (PrestaShop)   — API pull каждые 15 мин
  - Allegro (drogatrade) — OAuth2 + API pull каждые 15 мин

Railway deploy: добавь DB_CONNECT + API-ключи в env vars.
"""

import os
import io
import csv
import json
import asyncio
import hashlib
import hmac
import traceback
from contextlib import asynccontextmanager
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from base64 import b64encode
from decimal import Decimal

import asyncpg
import aiohttp
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

from fastapi import FastAPI, Request, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from auth import auth_router, get_current_user, require_user, has_perm, set_db_pool, init_auth_tables


# ============================================================
# Конфигурация
# ============================================================
TZ = ZoneInfo("Europe/Warsaw")
DSN = lambda: os.getenv("DB_CONNECT") or os.getenv("DATABASE_URL") or os.getenv("DATABASE_PRIVATE_URL")

# WooCommerce — dobraszklarnia.pl
WOO_URL     = os.getenv("WOO_URL", "https://dobraszklarnia.pl")
WOO_KEY     = os.getenv("WOO_KEY", "")       # Consumer Key
WOO_SECRET  = os.getenv("WOO_SECRET", "")     # Consumer Secret
WOO_WEBHOOK_SECRET = os.getenv("WOO_WEBHOOK_SECRET", "")  # Webhook secret для верификации

# PrestaShop — oteko.pl
PRESTA_OTEKO_URL = os.getenv("PRESTA_OTEKO_URL", "https://oteko.pl")
PRESTA_OTEKO_KEY = os.getenv("PRESTA_OTEKO_KEY", "")

# PrestaShop — cieplarnia.pl
PRESTA_CIEP_URL  = os.getenv("PRESTA_CIEP_URL", "https://cieplarnia.pl")
PRESTA_CIEP_KEY  = os.getenv("PRESTA_CIEP_KEY", "")

# Allegro — multi-account support
# Account 1: ALLEGRO_CLIENT_ID / ALLEGRO_CLIENT_SECRET / ALLEGRO_REFRESH_TOKEN / ALLEGRO_SELLER_NAME
# Account 2: ALLEGRO2_CLIENT_ID / ALLEGRO2_CLIENT_SECRET / ALLEGRO2_REFRESH_TOKEN / ALLEGRO2_SELLER_NAME
# ... up to ALLEGRO10_*
ALLEGRO_API_URL  = "https://api.allegro.pl"
ALLEGRO_AUTH_URL = "https://allegro.pl/auth/oauth/token"

def _load_allegro_accounts():
    """Load all Allegro accounts from env vars."""
    accounts = []
    # First account: ALLEGRO_* (no number)
    prefixes = [("ALLEGRO", "")] + [(f"ALLEGRO{i}", str(i)) for i in range(2, 11)]
    for prefix, suffix in prefixes:
        cid = os.getenv(f"{prefix}_CLIENT_ID", "")
        csecret = os.getenv(f"{prefix}_CLIENT_SECRET", "")
        rt = os.getenv(f"{prefix}_REFRESH_TOKEN", "")
        name = os.getenv(f"{prefix}_SELLER_NAME", f"allegro{suffix}" if suffix else "mantrade")
        if cid and csecret:
            accounts.append({
                "name": name,
                "client_id": cid,
                "client_secret": csecret,
                "refresh_token": rt,
                "prefix": prefix,
            })
    return accounts

ALLEGRO_ACCOUNTS = _load_allegro_accounts()

# In-memory token cache per account: { "account_name": {"access_token": ..., "expires_at": ..., "refresh_token": ...} }
_allegro_tokens = {}

# Bitrix24 (фаза 3 — пока закомментировано)
# BITRIX_WEBHOOK_URL = os.getenv("BITRIX_WEBHOOK_URL")


# ============================================================
# Database connection pool
# ============================================================
_pool = None

async def get_db_pool():
    global _pool
    if _pool is None or _pool._closed:
        _pool = await asyncpg.create_pool(
            dsn=DSN(),
            min_size=2,
            max_size=10,
            command_timeout=30,
        )
    return _pool


# ============================================================
# Lifespan — создание таблиц + запуск планировщика
# ============================================================
scheduler = AsyncIOScheduler(timezone=TZ)

@asynccontextmanager
async def lifespan(app: FastAPI):
    # --- Startup ---
    dsn = DSN()
    print(f"[startup] DSN present: {bool(dsn)}")
    if dsn:
        # Маскируем пароль для лога
        safe = dsn[:30] + "..." if len(dsn) > 30 else dsn
        print(f"[startup] DSN starts with: {safe}")
    else:
        # Выводим все env vars с DB/DATABASE в имени для диагностики
        db_vars = {k: v[:20]+"..." for k, v in os.environ.items() if "DB" in k.upper() or "DATABASE" in k.upper() or "POSTGRES" in k.upper() or "PG" in k.upper()}
        print(f"[startup] WARNING: No DSN found! DB-related env vars: {db_vars}")
        print(f"[startup] All env var names: {list(os.environ.keys())}")

    pool = await get_db_pool()
    set_db_pool(pool)  # share pool with auth module
    async with pool.acquire() as conn:
        # Auth tables (users, sessions)
        await init_auth_tables(conn)

        # Таблица лидов (из форм, звонков)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS leads (
                id            SERIAL PRIMARY KEY,
                name          VARCHAR(200),
                phone         VARCHAR(50),
                email         VARCHAR(200),
                site          VARCHAR(200),
                brand_slug    VARCHAR(100) DEFAULT 'unknown',
                form_type     VARCHAR(50)  DEFAULT 'form',
                utm_source    VARCHAR(200),
                utm_medium    VARCHAR(200),
                utm_campaign  VARCHAR(500),
                utm_term      VARCHAR(500),
                gclid         VARCHAR(500),
                landing_page  TEXT,
                campaign_id   VARCHAR(100),
                ad_id         VARCHAR(100),
                status        VARCHAR(50) DEFAULT 'new',
                device        VARCHAR(50),
                created_at    TIMESTAMP DEFAULT NOW(),
                updated_at    TIMESTAMP DEFAULT NOW()
            )
        """)

        # Таблица заказов — единая для всех магазинов
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id              SERIAL PRIMARY KEY,
                external_id     VARCHAR(100),
                source          VARCHAR(50) NOT NULL,
                status          VARCHAR(50) DEFAULT 'new',
                customer_name   VARCHAR(300),
                customer_email  VARCHAR(200),
                customer_phone  VARCHAR(50),
                customer_city   VARCHAR(200),
                customer_zip    VARCHAR(20),
                customer_address TEXT,
                total           NUMERIC(12,2) DEFAULT 0,
                currency        VARCHAR(10) DEFAULT 'PLN',
                items_count     INTEGER DEFAULT 0,
                items_json      TEXT,
                payment_method  VARCHAR(100),
                shipping_method VARCHAR(200),
                note            TEXT,
                external_created TIMESTAMP,
                created_at      TIMESTAMP DEFAULT NOW(),
                updated_at      TIMESTAMP DEFAULT NOW(),
                UNIQUE(source, external_id)
            )
        """)

        # Таблица клиентов — уникальные по email/phone
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                id              SERIAL PRIMARY KEY,
                name            VARCHAR(300),
                email           VARCHAR(200),
                phone           VARCHAR(50),
                city            VARCHAR(200),
                source          VARCHAR(50),
                orders_count    INTEGER DEFAULT 0,
                total_spent     NUMERIC(12,2) DEFAULT 0,
                first_order     TIMESTAMP,
                last_order      TIMESTAMP,
                created_at      TIMESTAMP DEFAULT NOW(),
                updated_at      TIMESTAMP DEFAULT NOW()
            )
        """)

        # Лог синхронизации — отслеживание процесса
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS sync_log (
                id          SERIAL PRIMARY KEY,
                source      VARCHAR(50),
                status      VARCHAR(20),
                orders_new  INTEGER DEFAULT 0,
                orders_upd  INTEGER DEFAULT 0,
                error       TEXT,
                started_at  TIMESTAMP DEFAULT NOW(),
                finished_at TIMESTAMP
            )
        """)

        # Лог экспортов — для отката
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS export_log (
                id           SERIAL PRIMARY KEY,
                user_email   VARCHAR(200),
                order_ids    INTEGER[],
                orders_count INTEGER DEFAULT 0,
                created_at   TIMESTAMP DEFAULT NOW(),
                undone       BOOLEAN DEFAULT FALSE,
                undone_at    TIMESTAMP
            )
        """)

        # App settings — key-value store (для Allegro refresh token и др.)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                key         VARCHAR(100) PRIMARY KEY,
                value       TEXT,
                updated_at  TIMESTAMP DEFAULT NOW()
            )
        """)

        # Auto-run migrations (add missing columns to existing tables)
        auto_migrations = [
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS product_type VARCHAR(100) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS configuration TEXT DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS shipping_cost NUMERIC(10,2) DEFAULT 0",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS seller_account VARCHAR(200) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS customer_comment TEXT DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS exported BOOLEAN DEFAULT FALSE",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS exported_at TIMESTAMP",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS is_priority BOOLEAN DEFAULT FALSE",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS invoice_nip VARCHAR(50) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS created_by_email VARCHAR(200) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS utm_source VARCHAR(200) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS utm_medium VARCHAR(200) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS utm_campaign VARCHAR(500) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS utm_content VARCHAR(500) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS utm_term VARCHAR(500) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS traffic_source_type VARCHAR(50) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS device_type VARCHAR(50) DEFAULT ''",
        ]
        for m in auto_migrations:
            try:
                await conn.execute(m)
            except Exception:
                pass
        print("[startup] auto-migrations applied")

    print("[startup] DB connected, all tables ready")

    # Фоновые задачи — синк магазинов
    scheduler.add_job(sync_woocommerce, CronTrigger(minute="*/15"), id="sync_woo", replace_existing=True)
    scheduler.add_job(sync_presta_oteko, CronTrigger(minute="*/15"), id="sync_oteko", replace_existing=True)
    scheduler.add_job(sync_presta_ciep, CronTrigger(minute="*/15"), id="sync_ciep", replace_existing=True)
    scheduler.add_job(sync_allegro, CronTrigger(minute="*/15"), id="sync_allegro", replace_existing=True)
    scheduler.start()
    allegro_names = [a["name"] for a in ALLEGRO_ACCOUNTS]
    print(f"[startup] scheduler started — syncing every 15 min (woo + oteko + ciep + allegro x{len(ALLEGRO_ACCOUNTS)}: {allegro_names})")

    yield

    # --- Shutdown ---
    scheduler.shutdown(wait=False)
    if _pool and not _pool._closed:
        await _pool.close()
    print("[shutdown] cleanup done")


# ============================================================
# FastAPI app
# ============================================================
app = FastAPI(lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

def _safe_val(v):
    """Конвертирует Decimal/datetime в JSON-совместимые типы."""
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.isoformat()
    return v

def _safe_dict(row):
    """Конвертирует asyncpg Record в JSON-safe dict."""
    return {k: _safe_val(v) for k, v in dict(row).items()}

# static — создаём папку если нет (Railway может не включить пустую папку)
import pathlib
_static_dir = pathlib.Path("static")
_static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="templates")
app.include_router(auth_router)


# ============================================================
# === ИНТЕГРАЦИИ: WooCommerce (dobraszklarnia.pl) ===
# ============================================================

def _woo_auth_header():
    """Basic Auth для WooCommerce REST API."""
    creds = b64encode(f"{WOO_KEY}:{WOO_SECRET}".encode()).decode()
    return {"Authorization": f"Basic {creds}"}


async def _woo_fetch_orders(session: aiohttp.ClientSession, page=1, per_page=50, after=None):
    """Получить страницу заказов из WooCommerce API."""
    params = {"page": page, "per_page": per_page, "orderby": "date", "order": "desc"}
    if after:
        params["after"] = after  # ISO 8601 формат
    url = f"{WOO_URL}/wp-json/wc/v3/orders"
    async with session.get(url, headers=_woo_auth_header(), params=params, ssl=True) as resp:
        if resp.status != 200:
            text = await resp.text()
            raise Exception(f"WooCommerce API error {resp.status}: {text[:300]}")
        return await resp.json()


async def _save_order(conn, source: str, order_data: dict):
    """Сохранить/обновить заказ в unified таблице orders."""
    ext_id = str(order_data["external_id"])

    existing = await conn.fetchval(
        "SELECT id FROM orders WHERE source=$1 AND external_id=$2",
        source, ext_id
    )

    if existing:
        await conn.execute("""
            UPDATE orders SET
                status=$1, customer_name=$2, customer_email=$3, customer_phone=$4,
                customer_city=$5, customer_zip=$6, customer_address=$7,
                total=$8, currency=$9, items_count=$10, items_json=$11,
                payment_method=$12, shipping_method=$13, note=$14,
                external_created=$15, shipping_cost=$18,
                seller_account=$19,
                customer_comment=CASE WHEN customer_comment IS NOT NULL AND customer_comment != '' THEN customer_comment ELSE $20 END,
                invoice_nip=$21,
                product_type=$22,
                utm_source=CASE WHEN $23 != '' THEN $23 ELSE utm_source END,
                utm_medium=CASE WHEN $24 != '' THEN $24 ELSE utm_medium END,
                utm_campaign=CASE WHEN $25 != '' THEN $25 ELSE utm_campaign END,
                utm_content=CASE WHEN $26 != '' THEN $26 ELSE utm_content END,
                utm_term=CASE WHEN $27 != '' THEN $27 ELSE utm_term END,
                traffic_source_type=CASE WHEN $28 != '' THEN $28 ELSE traffic_source_type END,
                device_type=CASE WHEN $29 != '' THEN $29 ELSE device_type END,
                updated_at=NOW()
            WHERE source=$16 AND external_id=$17
        """,
            order_data.get("status", ""),
            order_data.get("customer_name", ""),
            order_data.get("customer_email", ""),
            order_data.get("customer_phone", ""),
            order_data.get("customer_city", ""),
            order_data.get("customer_zip", ""),
            order_data.get("customer_address", ""),
            float(order_data.get("total", 0)),
            order_data.get("currency", "PLN"),
            int(order_data.get("items_count", 0)),
            order_data.get("items_json", "[]"),
            order_data.get("payment_method", ""),
            order_data.get("shipping_method", ""),
            order_data.get("note", ""),
            order_data.get("external_created"),
            source, ext_id,
            float(order_data.get("shipping_cost", 0)),
            order_data.get("seller_account", ""),
            order_data.get("customer_comment", ""),
            order_data.get("invoice_nip", ""),
            order_data.get("product_type", ""),
            order_data.get("utm_source", ""),
            order_data.get("utm_medium", ""),
            order_data.get("utm_campaign", ""),
            order_data.get("utm_content", ""),
            order_data.get("utm_term", ""),
            order_data.get("traffic_source_type", ""),
            order_data.get("device_type", ""),
        )
        return "updated"
    else:
        await conn.execute("""
            INSERT INTO orders (
                external_id, source, status, customer_name, customer_email,
                customer_phone, customer_city, customer_zip, customer_address,
                total, currency, items_count, items_json,
                payment_method, shipping_method, note, external_created,
                shipping_cost, seller_account, customer_comment, invoice_nip,
                product_type, utm_source, utm_medium, utm_campaign,
                utm_content, utm_term, traffic_source_type, device_type
            ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18,$19,$20,$21,$22,$23,$24,$25,$26,$27,$28,$29)
        """,
            ext_id, source,
            order_data.get("status", ""),
            order_data.get("customer_name", ""),
            order_data.get("customer_email", ""),
            order_data.get("customer_phone", ""),
            order_data.get("customer_city", ""),
            order_data.get("customer_zip", ""),
            order_data.get("customer_address", ""),
            float(order_data.get("total", 0)),
            order_data.get("currency", "PLN"),
            int(order_data.get("items_count", 0)),
            order_data.get("items_json", "[]"),
            order_data.get("payment_method", ""),
            order_data.get("shipping_method", ""),
            order_data.get("note", ""),
            order_data.get("external_created"),
            float(order_data.get("shipping_cost", 0)),
            order_data.get("seller_account", ""),
            order_data.get("customer_comment", ""),
            order_data.get("invoice_nip", ""),
            order_data.get("product_type", ""),
            order_data.get("utm_source", ""),
            order_data.get("utm_medium", ""),
            order_data.get("utm_campaign", ""),
            order_data.get("utm_content", ""),
            order_data.get("utm_term", ""),
            order_data.get("traffic_source_type", ""),
            order_data.get("device_type", ""),
        )
        return "new"


async def _upsert_customer(conn, source: str, name: str, email: str, phone: str, city: str,
                            total: float, order_date):
    """Обновить/создать клиента. Матчим по email (приоритет) или phone."""
    match_field = None
    match_value = None
    if email:
        match_field = "email"
        match_value = email
    elif phone:
        match_field = "phone"
        match_value = phone
    else:
        return

    existing = await conn.fetchrow(
        f"SELECT id, orders_count, total_spent FROM customers WHERE {match_field}=$1",
        match_value
    )

    if existing:
        await conn.execute("""
            UPDATE customers SET
                name = COALESCE(NULLIF($1, ''), name),
                city = COALESCE(NULLIF($2, ''), city),
                orders_count = orders_count + 1,
                total_spent = total_spent + $3,
                last_order = $4,
                updated_at = NOW()
            WHERE id = $5
        """, name, city, total, order_date, existing["id"])
    else:
        await conn.execute("""
            INSERT INTO customers (name, email, phone, city, source, orders_count, total_spent, first_order, last_order)
            VALUES ($1, $2, $3, $4, $5, 1, $6, $7, $7)
        """, name, email, phone, city, source, total, order_date)


import re as _re_nip


def _clean_product_display(name: str) -> str:
    """Strip leading product type keywords from display name.
    E.g. 'Poliwęglan komorowy 4,5x2,1 10mm' -> 'komorowy 4,5x2,1 10mm'
         'Szklarnia z poliwęglanu Mocna 3x6' -> 'Mocna 3x6'
         'EKO Szklarnia z poliwęglanu 3x8 4mm' -> '3x8 4mm'
    """
    if not name:
        return name
    import re
    # Remove leading type phrases (case-insensitive), including multi-word combos
    # "Szklarnia z poliwęglanu", "EKO Szklarnia z poliwęglanu", "Poliwęglan komorowy", etc.
    cleaned = re.sub(
        r'^(?:eko\s+)?(?:szklarni[ae]?\s+(?:z\s+poliw[eę]glanu?\s*)?|poliw[eę]glan\s+|cieplarni[ae]?\s+|теплица\s+|поликарбонат\s+)',
        '', name, flags=re.IGNORECASE
    )
    return cleaned.strip() or name


def _is_valid_nip(val: str) -> bool:
    """Check if a value looks like a real NIP/tax ID (must contain at least 5 digits)."""
    if not val:
        return False
    v = val.strip()
    if v.lower() in ("", "0", "-", "--", ".", "n/a", "brak", "none", "null"):
        return False
    # Must contain at least 5 digit characters
    digits = _re_nip.findall(r'\d', v)
    return len(digits) >= 5


def _detect_product_type(items: list) -> str:
    """Auto-detect product type from item names/SKUs.
    Each item is classified independently:
      - If item contains szklarnia keywords -> greenhouse (even if it also mentions poliweglan)
      - If item contains poliweglan keywords but NOT szklarnia -> standalone polycarbonate
    Then combine across all items in the order.
    Returns: 'Теплица', 'Поликарбонат', 'Теплица + Поликарбонат', or ''
    """
    has_szklarnia = False
    has_standalone_poliweglan = False
    szklarnia_kw = ("szklarnia", "szklarni", "теплиц", "cieplarni", "greenhouse", "tunel")
    poliweglan_kw = ("poliwęglan", "poliwegl", "поликарбонат", "polycarbonate", "komorow", "lity")
    szklarnia_sku_kw = ("szk", "lux", "eco", "eko")
    poliweglan_sku_kw = ("pc", "poly", "poliw")

    for item in items:
        name_lower = (item.get("full_name") or item.get("name") or "").lower()
        sku_lower = (item.get("sku") or "").lower()
        combined = name_lower + " " + sku_lower

        item_is_szklarnia = (
            any(kw in combined for kw in szklarnia_kw)
            or any(kw in sku_lower for kw in szklarnia_sku_kw)
        )
        item_is_poliweglan = (
            any(kw in combined for kw in poliweglan_kw)
            or any(kw in sku_lower for kw in poliweglan_sku_kw)
        )

        if item_is_szklarnia:
            # "Szklarnia z poliwęglanu" = greenhouse, not polycarbonate
            has_szklarnia = True
        elif item_is_poliweglan:
            # Standalone polycarbonate (no szklarnia keywords in this item)
            has_standalone_poliweglan = True

    if has_szklarnia and has_standalone_poliweglan:
        return "Теплица + Поликарбонат"
    if has_szklarnia:
        return "Теплица"
    if has_standalone_poliweglan:
        return "Поликарбонат"
    return ""


def _parse_woo_order(woo: dict) -> dict:
    """Маппинг WooCommerce order → наш формат."""
    billing = woo.get("billing", {})
    items = []
    for item in woo.get("line_items", []):
        product_name = item.get("name", "")
        # Append product attributes (Długość, Grubość, etc.) if not already in the name
        meta_parts = []
        for m in item.get("meta_data", []):
            key = m.get("display_key") or m.get("key", "")
            val = m.get("display_value") or m.get("value", "")
            if not key or not val:
                continue
            # Skip internal WooCommerce meta keys (start with _)
            if key.startswith("_"):
                continue
            # Only add if the value isn't already in the product name
            if str(val).lower() not in product_name.lower():
                meta_parts.append(f"{val}")
        if meta_parts:
            product_name = f"{product_name} ({', '.join(meta_parts)})"
        items.append({
            "name": product_name,
            "qty": item.get("quantity", 1),
            "price": item.get("total", "0"),
            "sku": item.get("sku", ""),
        })

    ext_created = None
    if woo.get("date_created"):
        try:
            ext_created = datetime.fromisoformat(woo["date_created"].replace("Z", "+00:00"))
        except Exception:
            ext_created = datetime.now(TZ)

    # Shipping cost
    ship_cost = 0.0
    try:
        ship_cost = float(woo.get("shipping_total", 0) or 0)
    except (ValueError, TypeError):
        ship_cost = 0.0

    # Customer comment
    customer_comment = woo.get("customer_note", "") or ""

    # NIP extraction — from WooCommerce billing fields and meta_data
    invoice_nip = ""
    woo_id = woo.get("id", "?")
    # 1) Direct billing field
    for bkey in ("nip", "company_nip", "nip_number"):
        bval = billing.get(bkey)
        if _is_valid_nip(str(bval or "")):
            invoice_nip = str(bval).strip()
            print(f"[woo-nip] order {woo_id}: found in billing.{bkey} = {invoice_nip}")
            break
    # 2) Search meta_data: specifically _billing_nip / billing_nip (the actual NIP value)
    if not invoice_nip:
        for m in woo.get("meta_data", []):
            key = (m.get("key") or "").lower().strip()
            val = str(m.get("value") or "").strip()
            if key in ("_billing_nip", "billing_nip", "_nip", "nip") and _is_valid_nip(val):
                invoice_nip = val
                print(f"[woo-nip] order {woo_id}: found in meta '{key}' = {invoice_nip}")
                break
    # 3) If still nothing, try vat_number keys
    if not invoice_nip:
        for m in woo.get("meta_data", []):
            key = (m.get("key") or "").lower().strip()
            val = str(m.get("value") or "").strip()
            if key in ("_vat_number", "vat_number", "_tax_id", "tax_id") and _is_valid_nip(val):
                invoice_nip = val
                print(f"[woo-nip] order {woo_id}: found in meta '{key}' = {invoice_nip}")
                break
    # UTM / order attribution — WooCommerce 8.5+ stores this in meta_data
    utm_source = ""
    utm_medium = ""
    utm_campaign = ""
    utm_content = ""
    utm_term = ""
    traffic_source_type = ""
    device_type = ""
    for m in woo.get("meta_data", []):
        key = (m.get("key") or "").lower().strip()
        val = str(m.get("value") or "").strip()
        if not val:
            continue
        if key in ("_wc_order_attribution_utm_source", "utm_source"):
            utm_source = val
        elif key in ("_wc_order_attribution_utm_medium", "utm_medium"):
            utm_medium = val
        elif key in ("_wc_order_attribution_utm_campaign", "utm_campaign"):
            utm_campaign = val
        elif key in ("_wc_order_attribution_utm_content", "utm_content"):
            utm_content = val
        elif key in ("_wc_order_attribution_utm_term", "utm_term"):
            utm_term = val
        elif key == "_wc_order_attribution_source_type":
            traffic_source_type = val
            # Fallback: if no utm_medium set, use source_type
            if not utm_medium and val in ("organic", "direct", "referral"):
                utm_medium = val
        elif key == "_wc_order_attribution_device_type":
            device_type = val
        elif key == "_wc_order_attribution_referrer" and not utm_source:
            # Extract source from referrer URL as last resort
            if "google" in val.lower():
                utm_source = "google"
            elif "facebook" in val.lower() or "fb." in val.lower():
                utm_source = "facebook"
            elif "bing" in val.lower():
                utm_source = "bing"
    # If source_type says "organic" and we have a referrer-based source but no medium
    if traffic_source_type == "organic" and utm_source and not utm_medium:
        utm_medium = "organic"
    # If source_type is "direct" and no source/medium at all
    if traffic_source_type == "direct" and not utm_source:
        utm_source = "(direct)"
        utm_medium = "(none)"

    # Prefer shipping (recipient) data over billing (payer) for name & address
    shipping = woo.get("shipping", {})
    recip_name = f"{shipping.get('first_name', '')} {shipping.get('last_name', '')}".strip()
    if not recip_name:
        recip_name = f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip()
    recip_city = shipping.get("city") or billing.get("city", "")
    recip_zip = shipping.get("postcode") or billing.get("postcode", "")
    recip_addr = f"{shipping.get('address_1', '')} {shipping.get('address_2', '')}".strip()
    if not recip_addr:
        recip_addr = f"{billing.get('address_1', '')} {billing.get('address_2', '')}".strip()
    # Phone: shipping may have it (WC 8.3+), fall back to billing
    recip_phone = shipping.get("phone") or billing.get("phone", "")

    return {
        "external_id": str(woo["id"]),
        "status": woo.get("status", "unknown"),
        "customer_name": recip_name,
        "customer_email": billing.get("email", ""),
        "customer_phone": recip_phone,
        "customer_city": recip_city,
        "customer_zip": recip_zip,
        "customer_address": recip_addr,
        "total": float(woo.get("total", 0)),
        "currency": woo.get("currency", "PLN"),
        "items_count": len(items),
        "items_json": json.dumps(items, ensure_ascii=False),
        "payment_method": woo.get("payment_method_title", ""),
        "shipping_method": ", ".join(s.get("method_title", "") for s in woo.get("shipping_lines", [])),
        "note": woo.get("customer_note", ""),
        "external_created": ext_created,
        "shipping_cost": ship_cost,
        "customer_comment": customer_comment,
        "seller_account": "",
        "invoice_nip": invoice_nip,
        "product_type": _detect_product_type(items),
        "utm_source": utm_source,
        "utm_medium": utm_medium,
        "utm_campaign": utm_campaign,
        "utm_content": utm_content,
        "utm_term": utm_term,
        "traffic_source_type": traffic_source_type,
        "device_type": device_type,
    }


async def sync_woocommerce(full=False):
    """Фоновая задача: синк заказов из WooCommerce (dobraszklarnia.pl)."""
    if not WOO_KEY or not WOO_SECRET:
        return

    source = "dobraszklarnia"
    pool = await get_db_pool()

    # Определяем с какой даты синкать (последний заказ + запас)
    after = None
    if not full:
        async with pool.acquire() as conn:
            last = await conn.fetchval(
                "SELECT MAX(external_created) FROM orders WHERE source=$1", source
            )
        if last:
            after = (last - timedelta(hours=2)).isoformat()

    log_id = None
    async with pool.acquire() as conn:
        log_id = await conn.fetchval(
            "INSERT INTO sync_log (source, status) VALUES ($1, 'running') RETURNING id", source
        )

    new_count = 0
    upd_count = 0

    try:
        async with aiohttp.ClientSession() as session:
            page = 1
            while True:
                orders = await _woo_fetch_orders(session, page=page, per_page=50, after=after)
                if not orders:
                    break

                async with pool.acquire() as conn:
                    for woo_order in orders:
                        parsed = _parse_woo_order(woo_order)
                        result = await _save_order(conn, source, parsed)
                        if result == "new":
                            new_count += 1
                            # Upsert customer только для новых
                            await _upsert_customer(
                                conn, source,
                                parsed["customer_name"], parsed["customer_email"],
                                parsed["customer_phone"], parsed["customer_city"],
                                parsed["total"], parsed["external_created"]
                            )
                        else:
                            upd_count += 1

                if len(orders) < 50:
                    break
                page += 1

        async with pool.acquire() as conn:
            await conn.execute(
                "UPDATE sync_log SET status='ok', orders_new=$1, orders_upd=$2, finished_at=NOW() WHERE id=$3",
                new_count, upd_count, log_id
            )
        # Count orders with NIP in DB after sync
        async with pool.acquire() as conn:
            nip_count = await conn.fetchval(
                "SELECT COUNT(*) FROM orders WHERE source=$1 AND invoice_nip IS NOT NULL AND invoice_nip != ''",
                source
            )
        print(f"[sync-woo] done: +{new_count} new, ~{upd_count} updated, {nip_count} orders have NIP in DB")

    except Exception as e:
        print(f"[sync-woo] ERROR: {e}")
        traceback.print_exc()
        async with pool.acquire() as conn:
            await conn.execute(
                "UPDATE sync_log SET status='error', error=$1, finished_at=NOW() WHERE id=$2",
                str(e)[:500], log_id
            )


# ============================================================
# === ИНТЕГРАЦИИ: PrestaShop (oteko.pl + cieplarnia.pl) ===
# ============================================================

async def _presta_fetch_orders(session: aiohttp.ClientSession, base_url: str, api_key: str,
                                page=0, limit=50, since_id=None):
    """Получить заказы из PrestaShop Webservice API."""
    url = f"{base_url}/api/orders"
    params = {
        "output_format": "JSON",
        "display": "full",
        "limit": f"{page * limit},{limit}",
        "sort": "[id_DESC]",
    }
    if since_id:
        params["filter[id]"] = f"[{since_id},999999999]"

    auth = aiohttp.BasicAuth(api_key, "")  # PrestaShop: key как логин, пустой пароль
    full_url = f"{url}?{'&'.join(f'{k}={v}' for k,v in params.items())}"
    print(f"[presta] GET {full_url} (key={api_key[:8]}...)")
    async with session.get(url, params=params, auth=auth, ssl=True) as resp:
        print(f"[presta] response status={resp.status}, url={resp.url}")
        if resp.status == 401:
            text = await resp.text()
            raise Exception(f"PrestaShop API 401 Unauthorized: {text[:300]}")
        if resp.status != 200:
            text = await resp.text()
            raise Exception(f"PrestaShop API error {resp.status}: {text[:300]}")
        data = await resp.json()
        return data.get("orders", [])


async def _presta_fetch_customer(session: aiohttp.ClientSession, base_url: str, api_key: str,
                                  customer_id):
    """Получить данные клиента по ID."""
    if not customer_id:
        return {}
    url = f"{base_url}/api/customers/{customer_id}"
    params = {"output_format": "JSON"}
    auth = aiohttp.BasicAuth(api_key, "")
    try:
        async with session.get(url, params=params, auth=auth, ssl=True) as resp:
            if resp.status != 200:
                return {}
            data = await resp.json()
            return data.get("customer", {})
    except Exception:
        return {}


async def _presta_fetch_address(session: aiohttp.ClientSession, base_url: str, api_key: str,
                                 address_id):
    """Получить адрес по ID."""
    if not address_id:
        return {}
    url = f"{base_url}/api/addresses/{address_id}"
    params = {"output_format": "JSON"}
    auth = aiohttp.BasicAuth(api_key, "")
    try:
        async with session.get(url, params=params, auth=auth, ssl=True) as resp:
            if resp.status != 200:
                return {}
            data = await resp.json()
            return data.get("address", {})
    except Exception:
        return {}


# Маппинг статусов PrestaShop (стандартные ID → текст)
PRESTA_STATUS_MAP = {
    "1": "awaiting_check",
    "2": "payment_accepted",
    "3": "processing",
    "4": "shipped",
    "5": "delivered",
    "6": "cancelled",
    "7": "refunded",
    "8": "payment_error",
    "9": "order_pending_paid",
    "10": "awaiting_bankwire",
    "11": "payment_received",
    "12": "order_pending_unpaid",
    "13": "awaiting_cod",
    "14": "awaiting_payment",
    "15": "partial_refund",
    "16": "partial_payment",
    "17": "authorized_transfer",
    "18": "awaiting_przelewy24",
    "19": "przelewy24_accepted",
    "20": "payu_started",
    "21": "payu_awaiting",
    "22": "payu_cancelled",
}

# ---- Все статусы на польском ----
STATUS_PL = {
    # WooCommerce / general
    "pending": "Oczekujące",
    "pending_payment": "Oczekiwanie na płatność",
    "processing": "W realizacji",
    "shipped": "Wysłane",
    "delivered": "Dostarczone",
    "completed": "Zrealizowane",
    "cancelled": "Anulowane",
    "refunded": "Zwrot",
    "failed": "Nieudane",
    "on-hold": "Wstrzymane",
    "on_hold": "Wstrzymane",
    "trash": "Usunięte",
    # PrestaShop
    "payment_accepted": "Płatność zaakceptowana",
    "payment_error": "Błąd płatności",
    "payment_received": "Płatność przyjęta",
    "awaiting_bankwire": "Oczekiwanie na przelew",
    "awaiting_check": "Oczekiwanie na płatność czekiem",
    "awaiting_cod": "Oczekiwanie na płatność przy odbiorze",
    "awaiting_payment": "Oczekiwanie na płatność",
    "order_pending_paid": "Zamówienie oczekujące (opłacone)",
    "order_pending_unpaid": "Zamówienie oczekujące (nieopłacone)",
    "partial_refund": "Częściowy zwrot",
    "partial_payment": "Częściowa płatność",
    "authorized_transfer": "Autoryzacja — transfer przez sklep",
    "awaiting_przelewy24": "Oczekiwanie na Przelewy24",
    "przelewy24_accepted": "Płatność Przelewy24 przyjęta",
    "payu_started": "Płatność PayU rozpoczęta",
    "payu_awaiting": "PayU oczekuje na odbiór",
    "payu_cancelled": "PayU anulowana",
    # Allegro
    "bought": "Kupione",
    "filled_in": "Wypełnione",
    "ready_for_processing": "Gotowe do realizacji",
}

def _status_pl(status: str) -> str:
    """Zwraca polską nazwę statusu."""
    if not status:
        return "—"
    return STATUS_PL.get(status, STATUS_PL.get(status.lower(), status))


def _parse_presta_order(order: dict, customer: dict, address: dict) -> dict:
    """Маппинг PrestaShop order → наш формат."""
    import re as _re
    items = []
    order_rows = order.get("associations", {}).get("order_rows", [])
    if isinstance(order_rows, list):
        for row in order_rows:
            raw_name = row.get("product_name", "")
            # Clean PrestaShop product name: keep only Szerokość and Długość from params
            # Use rsplit to find the LAST '(' — avoids matching e.g. "(mleczny)" in product name
            clean_name = raw_name
            if ' - ' in raw_name and '(' in raw_name:
                last_paren = raw_name.rfind('(')
                if last_paren > 0 and raw_name.endswith(')'):
                    base_name = raw_name[:last_paren].strip()
                    params_str = raw_name[last_paren+1:-1]
                    kept = []
                    for param in params_str.split(' - '):
                        param = param.strip()
                        if param.lower().startswith('szeroko') or param.lower().startswith('długo'):
                            kept.append(param)
                    if kept:
                        clean_name = base_name + ' (' + ' - '.join(kept) + ')'
                    else:
                        clean_name = base_name
            items.append({
                "name": clean_name,
                "full_name": raw_name,
                "qty": int(row.get("product_quantity", 1)),
                "price": row.get("product_price", "0"),
                "sku": row.get("product_reference", ""),
            })

    ext_created = None
    if order.get("date_add"):
        try:
            ext_created = datetime.fromisoformat(order["date_add"])
        except Exception:
            ext_created = datetime.now(TZ)

    status_id = str(order.get("current_state", ""))
    status = PRESTA_STATUS_MAP.get(status_id, f"state_{status_id}")

    # Prefer delivery address name (recipient) over customer (account holder)
    cust_name = f"{address.get('firstname', '')} {address.get('lastname', '')}".strip()
    if not cust_name:
        cust_name = f"{customer.get('firstname', '')} {customer.get('lastname', '')}".strip()

    # Shipping cost
    ship_cost = 0.0
    try:
        total_shipping = float(order.get("total_shipping", 0) or 0)
        total_shipping_tax = float(order.get("total_shipping_tax_incl", 0) or 0)
        ship_cost = total_shipping_tax if total_shipping_tax > 0 else total_shipping
    except (ValueError, TypeError):
        ship_cost = 0.0

    # Customer comment / note
    customer_comment = ""
    note_text = order.get("note", "") or ""
    if note_text:
        customer_comment = note_text

    # NIP extraction from PrestaShop address/customer
    invoice_nip = ""
    # Check address fields
    for nip_field in ("vat_number", "dni", "nip", "company_nip"):
        val = address.get(nip_field)
        if _is_valid_nip(str(val or "")):
            invoice_nip = str(val).strip()
            break
    # Check customer fields
    if not invoice_nip:
        for nip_field in ("vat_number", "dni", "nip", "siret"):
            val = customer.get(nip_field)
            if _is_valid_nip(str(val or "")):
                invoice_nip = str(val).strip()
                break
    # Try company field with NIP regex
    if not invoice_nip and address.get("company") and str(address.get("company", "")).strip():
        import re as _re2
        nip_match = _re2.search(r'(?:NIP|nip)[:\s]*(\d[\d\s-]{8,}\d)', str(address.get("company", "")))
        if nip_match:
            invoice_nip = nip_match.group(1).replace(" ", "").replace("-", "")
    # Debug logging
    order_id = order.get("id", "?")
    if invoice_nip:
        print(f"[presta-nip] order {order_id}: NIP={invoice_nip}")
    else:
        # Log available address/customer fields for debug
        addr_nip_fields = {k: address.get(k) for k in ("vat_number", "dni", "company", "nip") if address.get(k) and str(address.get(k)).strip() and str(address.get(k)).strip() != "0"}
        cust_nip_fields = {k: customer.get(k) for k in ("vat_number", "dni", "siret", "company") if customer.get(k) and str(customer.get(k)).strip() and str(customer.get(k)).strip() != "0"}
        if addr_nip_fields or cust_nip_fields:
            print(f"[presta-nip] order {order_id}: addr={addr_nip_fields}, cust={cust_nip_fields}")

    return {
        "external_id": str(order.get("id", "")),
        "status": status,
        "customer_name": cust_name,
        "customer_email": customer.get("email", ""),
        "customer_phone": address.get("phone", "") or address.get("phone_mobile", ""),
        "customer_city": address.get("city", ""),
        "customer_zip": address.get("postcode", ""),
        "customer_address": f"{address.get('address1', '')} {address.get('address2', '')}".strip(),
        "total": float(order.get("total_paid", 0)),
        "currency": order.get("id_currency", "PLN"),
        "items_count": len(items),
        "items_json": json.dumps(items, ensure_ascii=False),
        "payment_method": order.get("payment", ""),
        "shipping_method": "",
        "note": "",
        "external_created": ext_created,
        "shipping_cost": ship_cost,
        "customer_comment": customer_comment,
        "seller_account": "",
        "invoice_nip": invoice_nip,
        "product_type": _detect_product_type(items),
        "utm_source": "",
        "utm_medium": "",
        "utm_campaign": "",
        "utm_content": "",
        "utm_term": "",
        "traffic_source_type": "",
        "device_type": "",
    }


async def _sync_prestashop(source: str, base_url: str, api_key: str, full=False):
    """Универсальная функция синка PrestaShop магазина."""
    if not api_key:
        return

    pool = await get_db_pool()

    # Определяем с какого ID синкать (incremental), full=True — всё с начала
    since_id = None
    if not full:
        async with pool.acquire() as conn:
            last_id = await conn.fetchval(
                "SELECT MAX(CAST(external_id AS INTEGER)) FROM orders WHERE source=$1 AND external_id ~ '^[0-9]+$'", source
            )
        if last_id:
            since_id = max(1, last_id - 10)  # small overlap for safety

    log_id = None
    async with pool.acquire() as conn:
        log_id = await conn.fetchval(
            "INSERT INTO sync_log (source, status) VALUES ($1, 'running') RETURNING id", source
        )

    new_count = 0
    upd_count = 0

    try:
        async with aiohttp.ClientSession() as session:
            page = 0
            while True:
                orders = await _presta_fetch_orders(
                    session, base_url, api_key,
                    page=page, limit=50, since_id=since_id
                )
                if not orders:
                    break

                async with pool.acquire() as conn:
                    for ps_order in orders:
                        try:
                            # Подтягиваем клиента и адрес
                            customer = await _presta_fetch_customer(
                                session, base_url, api_key, ps_order.get("id_customer")
                            )
                            address = await _presta_fetch_address(
                                session, base_url, api_key, ps_order.get("id_address_delivery")
                            )

                            parsed = _parse_presta_order(ps_order, customer, address)
                            result = await _save_order(conn, source, parsed)
                            if result == "new":
                                new_count += 1
                                await _upsert_customer(
                                    conn, source,
                                    parsed["customer_name"], parsed["customer_email"],
                                    parsed["customer_phone"], parsed["customer_city"],
                                    parsed["total"], parsed["external_created"]
                                )
                            else:
                                upd_count += 1
                        except Exception as e_order:
                            order_id = ps_order.get("id", "?")
                            print(f"[sync-{source}] SKIP order {order_id}: {e_order}")

                if len(orders) < 50:
                    break
                page += 1

        async with pool.acquire() as conn:
            await conn.execute(
                "UPDATE sync_log SET status='ok', orders_new=$1, orders_upd=$2, finished_at=NOW() WHERE id=$3",
                new_count, upd_count, log_id
            )
        print(f"[sync-{source}] done: +{new_count} new, ~{upd_count} updated")

    except Exception as e:
        print(f"[sync-{source}] ERROR: {e}")
        traceback.print_exc()
        async with pool.acquire() as conn:
            await conn.execute(
                "UPDATE sync_log SET status='error', error=$1, finished_at=NOW() WHERE id=$2",
                str(e)[:500], log_id
            )


async def sync_presta_oteko(full=False):
    await _sync_prestashop("oteko", PRESTA_OTEKO_URL, PRESTA_OTEKO_KEY, full=full)

async def sync_presta_ciep(full=False):
    await _sync_prestashop("cieplarnia", PRESTA_CIEP_URL, PRESTA_CIEP_KEY, full=full)


# ============================================================
# === ИНТЕГРАЦИИ: Allegro (drogatrade) ===
# ============================================================

import time as _time

async def _allegro_db_get_refresh_token(account_name: str):
    """Read the latest refresh token for a specific account from DB."""
    try:
        pool = await get_db_pool()
        db_key = f"allegro_refresh_token_{account_name}"
        async with pool.acquire() as conn:
            # Try new key format first, fallback to old format for backward compat
            row = await conn.fetchval(
                "SELECT value FROM app_settings WHERE key=$1", db_key
            )
            if not row and account_name == "mantrade":
                row = await conn.fetchval(
                    "SELECT value FROM app_settings WHERE key='allegro_refresh_token'"
                )
            return row
    except Exception:
        return None


async def _allegro_db_save_refresh_token(account_name: str, token: str):
    """Save new refresh token for a specific account to DB."""
    try:
        pool = await get_db_pool()
        db_key = f"allegro_refresh_token_{account_name}"
        async with pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO app_settings (key, value, updated_at)
                VALUES ($1, $2, NOW())
                ON CONFLICT (key) DO UPDATE SET value=$2, updated_at=NOW()
            """, db_key, token)
    except Exception as e:
        print(f"[allegro:{account_name}] WARNING: could not save refresh_token to DB: {e}")


async def _allegro_get_token(session: aiohttp.ClientSession, account: dict) -> str:
    """Get a valid Allegro access token for a specific account, refreshing if needed."""
    global _allegro_tokens

    name = account["name"]
    now = _time.time()

    # Check cache
    cached = _allegro_tokens.get(name, {})
    if cached.get("access_token") and cached.get("expires_at", 0) > now + 60:
        return cached["access_token"]

    # Priority: memory cache → DB → env var
    current_rt = (
        cached.get("refresh_token")
        or await _allegro_db_get_refresh_token(name)
        or account.get("refresh_token", "")
    )

    if not current_rt:
        raise Exception(f"No Allegro refresh token for account '{name}' (check env var {account['prefix']}_REFRESH_TOKEN)")

    # Refresh the token
    auth = aiohttp.BasicAuth(account["client_id"], account["client_secret"])
    data = {
        "grant_type": "refresh_token",
        "refresh_token": current_rt,
    }
    async with session.post(ALLEGRO_AUTH_URL, data=data, auth=auth) as resp:
        if resp.status != 200:
            text = await resp.text()
            raise Exception(f"Allegro token refresh failed for '{name}' {resp.status}: {text[:300]}")
        body = await resp.json()

    _allegro_tokens[name] = {
        "access_token": body["access_token"],
        "expires_at": now + body.get("expires_in", 3600) - 60,
    }

    # Store new refresh_token in memory + DB (Allegro rotates them on each use!)
    new_rt = body.get("refresh_token")
    if new_rt:
        _allegro_tokens[name]["refresh_token"] = new_rt
        await _allegro_db_save_refresh_token(name, new_rt)
        if new_rt != current_rt:
            print(f"[allegro:{name}] refresh_token rotated and saved to DB")

    return _allegro_tokens[name]["access_token"]


async def _allegro_fetch_orders(session: aiohttp.ClientSession, token: str,
                                 offset=0, limit=100, updated_after=None):
    """Fetch checkout-forms (orders) from Allegro REST API."""
    url = f"{ALLEGRO_API_URL}/order/checkout-forms"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.allegro.public.v1+json",
    }
    params = {
        "offset": offset,
        "limit": min(limit, 100),
        "sort": "-updatedAt",
    }
    if updated_after:
        # Allegro requires ISO 8601 with timezone, e.g. 2026-04-14T10:00:00.000Z
        if isinstance(updated_after, str) and "T" not in updated_after:
            updated_after = updated_after + "T00:00:00.000Z"
        elif isinstance(updated_after, str) and not updated_after.endswith("Z") and "+" not in updated_after:
            updated_after = updated_after + "Z"
        params["updatedAt.gte"] = updated_after

    async with session.get(url, headers=headers, params=params) as resp:
        if resp.status == 401:
            raise Exception("Allegro API: token expired or invalid")
        if resp.status != 200:
            text = await resp.text()
            raise Exception(f"Allegro API error {resp.status}: {text[:300]}")
        data = await resp.json()
        forms = data.get("checkoutForms", [])
        # Allegro: "count" = items on this page, "totalCount" = total matching
        total_count = data.get("totalCount") or data.get("count") or len(forms)
        print(f"[allegro-api] response keys={list(data.keys())}, count={data.get('count')}, totalCount={data.get('totalCount')}, forms={len(forms)}")
        return forms, total_count


def _parse_allegro_order(form: dict, seller_name: str = "drogatrade") -> dict:
    """Map Allegro checkout-form → our unified order format."""
    buyer = form.get("buyer") or {}
    delivery = form.get("delivery") or {}
    address = delivery.get("address") or {}
    payment = form.get("payment") or {}
    summary = form.get("summary") or {}

    # Build recipient name (delivery address = who receives the package)
    recipient_name = ""
    if address.get("firstName") or address.get("lastName"):
        recipient_name = f"{address.get('firstName', '')} {address.get('lastName', '')}".strip()
    # Fallback to buyer name if delivery address has no name
    if not recipient_name:
        recipient_name = f"{buyer.get('firstName', '')} {buyer.get('lastName', '')}".strip()

    # Items — use cleaned full_name as display name, keep sygnatura as SKU
    items = []
    for li in form.get("lineItems", []):
        offer = li.get("offer") or {}
        ext = offer.get("external") or {}
        sygnatura = ext.get("id", "")
        full_name = offer.get("name", "")
        display_name = _clean_product_display(full_name) if full_name else sygnatura
        items.append({
            "name": display_name,
            "full_name": full_name,
            "qty": int(li.get("quantity", 1)),
            "price": str(li.get("price", {}).get("amount", "0")),
            "sku": sygnatura or str(offer.get("id", "")),
        })

    # Total
    total_amount = 0.0
    if summary.get("totalToPay"):
        total_amount = float(summary["totalToPay"].get("amount", 0))

    # Currency
    currency = "PLN"
    if summary.get("totalToPay", {}).get("currency"):
        currency = summary["totalToPay"]["currency"]

    # Created date — strip timezone to match our naive TIMESTAMP column
    ext_created = None
    bought_at = form.get("updatedAt") or form.get("createdAt")
    if bought_at:
        try:
            dt = datetime.fromisoformat(bought_at.replace("Z", "+00:00"))
            ext_created = dt.replace(tzinfo=None)
        except Exception:
            ext_created = datetime.now(TZ).replace(tzinfo=None)

    # Status mapping
    status_raw = form.get("status", "UNKNOWN")
    allegro_status_map = {
        "BOUGHT": "processing",
        "FILLED_IN": "processing",
        "READY_FOR_PROCESSING": "processing",
        "CANCELLED": "cancelled",
    }
    status = allegro_status_map.get(status_raw, status_raw.lower())

    # Check fulfillment status for shipped/delivered
    fulfillment = form.get("fulfillment") or {}
    fulfill_status = (fulfillment.get("status") or "").upper()
    if fulfill_status == "SENT":
        status = "shipped"
    elif fulfill_status == "PICKED_UP":
        status = "delivered"

    # Payment type → Polish label
    allegro_payment_map = {
        "CASH_ON_DELIVERY": "Płatność przy odbiorze",
        "ONLINE": "Opłacone",
        "TRANSFER": "Przelew",
        "CARD": "Karta",
        "INSTALLMENTS": "Raty",
        "P24": "Przelewy24",
        "SPLIT_PAYMENT": "Podzielona płatność",
    }
    pay_type_raw = (payment.get("type") or "").upper()
    payment_label = allegro_payment_map.get(pay_type_raw, pay_type_raw)

    # Payment status can override order status
    pay_status = pay_type_raw.lower()
    paid_amount_obj = payment.get("paidAmount") or {}
    if pay_status == "paid" or pay_type_raw == "ONLINE" or paid_amount_obj.get("amount"):
        paid_amt = float(paid_amount_obj.get("amount", 0) or 0)
        if paid_amt >= total_amount and total_amount > 0:
            if status == "processing":
                status = "payment_accepted"

    # Shipping cost
    ship_cost = 0.0
    delivery_cost = delivery.get("cost") or {}
    if delivery_cost.get("amount"):
        ship_cost = float(delivery_cost["amount"])

    # Seller comment / message to seller + NIP
    msg_to_seller = form.get("messageToSeller", "") or ""
    invoice = form.get("invoice") or {}
    invoice_nip = ""
    form_id = form.get("id", "?")
    # Try invoice.company.taxId (old format)
    if invoice.get("company") and invoice["company"].get("taxId"):
        invoice_nip = invoice["company"]["taxId"]
    # Try invoice.address.company.taxId (current Allegro API format)
    if not invoice_nip:
        inv_addr = invoice.get("address") or {}
        inv_company = inv_addr.get("company") or {}
        if inv_company.get("taxId"):
            invoice_nip = inv_company["taxId"]
        # Also check ids array: [{"type": "PL_NIP", "value": "..."}]
        if not invoice_nip and inv_company.get("ids"):
            for id_obj in inv_company["ids"]:
                if id_obj.get("type") == "PL_NIP" and id_obj.get("value"):
                    invoice_nip = id_obj["value"]
                    break
    if invoice_nip:
        print(f"[allegro-nip] order {form_id}: NIP={invoice_nip}")
    nip_label = f"NIP: {invoice_nip}" if invoice_nip else ""
    customer_comment = msg_to_seller
    if nip_label:
        customer_comment = f"{msg_to_seller}\n{nip_label}".strip() if msg_to_seller else nip_label

    # Seller account name
    seller_account = seller_name

    return {
        "external_id": str(form.get("id", "")),
        "status": status,
        "customer_name": recipient_name,
        "customer_email": buyer.get("email", ""),
        "customer_phone": address.get("phoneNumber", "") or buyer.get("phoneNumber", ""),
        "customer_city": address.get("city", ""),
        "customer_zip": address.get("zipCode", "") or address.get("postCode", ""),
        "customer_address": f"{address.get('street', '')}".strip(),
        "total": total_amount,
        "currency": currency,
        "items_count": len(items),
        "items_json": json.dumps(items, ensure_ascii=False),
        "shipping_cost": ship_cost,
        "customer_comment": customer_comment,
        "invoice_nip": invoice_nip,
        "product_type": _detect_product_type(items),
        "seller_account": seller_account,
        "payment_method": payment_label,
        "shipping_method": delivery.get("method", {}).get("name", ""),
        "note": form.get("messageToSeller", ""),
        "external_created": ext_created,
        "utm_source": "allegro",
        "utm_medium": "marketplace",
        "utm_campaign": "",
        "utm_content": "",
        "utm_term": "",
        "traffic_source_type": "marketplace",
        "device_type": "",
    }


async def _sync_allegro_account(account: dict, full=False):
    """Sync orders from a single Allegro account."""
    name = account["name"]
    source = "allegro"
    pool = await get_db_pool()

    # Find last sync date for this seller account
    updated_after = None
    if not full:
        async with pool.acquire() as conn:
            last = await conn.fetchval(
                "SELECT MAX(external_created) FROM orders WHERE source=$1 AND seller_account=$2",
                source, name
            )
            order_count = await conn.fetchval(
                "SELECT COUNT(*) FROM orders WHERE source=$1 AND seller_account=$2",
                source, name
            )
        if last and order_count and order_count > 10:
            # Use 7-day overlap to catch status changes (cancellations, etc.)
            updated_after = (last - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S.000Z")

    if full:
        print(f"[sync-allegro:{name}] FULL resync — fetching all orders")

    log_id = None
    async with pool.acquire() as conn:
        log_id = await conn.fetchval(
            "INSERT INTO sync_log (source, status) VALUES ($1, 'running') RETURNING id",
            f"allegro:{name}"
        )

    new_count = 0
    upd_count = 0
    skip_count = 0

    try:
        async with aiohttp.ClientSession() as session:
            token = await _allegro_get_token(session, account)

            if full:
                now = datetime.utcnow()
                date_ranges = []
                for m in range(4):
                    range_end = now - timedelta(days=90 * m)
                    range_start = now - timedelta(days=90 * (m + 1))
                    date_ranges.append((
                        range_start.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                        range_end.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                    ))
            else:
                date_ranges = [(updated_after, None)]

            for dr_start, dr_end in date_ranges:
                offset = 0
                while True:
                    forms, total = await _allegro_fetch_orders(
                        session, token,
                        offset=offset, limit=100, updated_after=dr_start
                    )
                    if offset == 0:
                        print(f"[sync-allegro:{name}] offset=0, total={total}, fetched={len(forms)}, range={dr_start}..{dr_end}")
                    if not forms:
                        break

                    async with pool.acquire() as conn:
                        for form in forms:
                            try:
                                parsed = _parse_allegro_order(form, seller_name=name)
                                result = await _save_order(conn, source, parsed)
                                if result == "new":
                                    new_count += 1
                                    await _upsert_customer(
                                        conn, source,
                                        parsed["customer_name"], parsed["customer_email"],
                                        parsed["customer_phone"], parsed["customer_city"],
                                        parsed["total"], parsed["external_created"]
                                    )
                                else:
                                    upd_count += 1
                            except Exception as e_order:
                                skip_count += 1
                                order_id = form.get("id", "?")
                                print(f"[sync-allegro:{name}] SKIP order {order_id}: {e_order}")

                    offset += len(forms)
                    if offset >= total or len(forms) < 100:
                        break

        async with pool.acquire() as conn:
            await conn.execute(
                "UPDATE sync_log SET status='ok', orders_new=$1, orders_upd=$2, finished_at=NOW() WHERE id=$3",
                new_count, upd_count, log_id
            )
        # Count allegro orders with NIP
        async with pool.acquire() as conn:
            nip_count = await conn.fetchval(
                "SELECT COUNT(*) FROM orders WHERE source='allegro' AND invoice_nip IS NOT NULL AND invoice_nip != ''"
            )
        print(f"[sync-allegro:{name}] done: +{new_count} new, ~{upd_count} updated, skipped={skip_count}, allegro NIP in DB={nip_count}")

    except Exception as e:
        print(f"[sync-allegro:{name}] ERROR: {e}")
        traceback.print_exc()
        async with pool.acquire() as conn:
            await conn.execute(
                "UPDATE sync_log SET status='error', error=$1, finished_at=NOW() WHERE id=$2",
                str(e)[:500], log_id
            )


async def sync_allegro(full=False):
    """Sync orders from ALL configured Allegro accounts."""
    if not ALLEGRO_ACCOUNTS:
        return
    for account in ALLEGRO_ACCOUNTS:
        try:
            await _sync_allegro_account(account, full=full)
        except Exception as e:
            print(f"[sync-allegro:{account['name']}] FATAL: {e}")
            traceback.print_exc()


# ============================================================
# === СТРАНИЦЫ ===
# ============================================================

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    user = await get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    # If user has no dashboard permission — redirect to orders (if they have it) or show empty
    if not has_perm(user, "dashboard"):
        if has_perm(user, "orders"):
            return RedirectResponse("/orders", status_code=302)
        return templates.TemplateResponse(
            request=request, name="no_access.html", context={"user": user}
        )
    pool = await get_db_pool()

    async with pool.acquire() as conn:
        # Только последний синк — всё остальное тянется динамически через /api/dashboard/stats
        last_syncs = await conn.fetch("""
            SELECT DISTINCT ON (source) source, status, orders_new, orders_upd, finished_at
            FROM sync_log ORDER BY source, finished_at DESC NULLS LAST
        """)

    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "user": user,
            "last_syncs": last_syncs,
        },
    )


@app.get("/orders", response_class=HTMLResponse)
async def orders_page(request: Request):
    user = await get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not has_perm(user, "orders"):
        return RedirectResponse("/", status_code=302)
    return templates.TemplateResponse(
        request=request,
        name="orders.html",
        context={"user": user},
    )


# ============================================================
# === API: Статистика ===
# ============================================================

@app.get("/api/stats")
async def api_stats(
    request: Request,
    date_from: str = Query(None),
    date_to:   str = Query(None),
    source:    str = Query(None),
    seller:    str = Query(None),
):
    user = await get_current_user(request)
    if not user or not (has_perm(user, "dashboard") or has_perm(user, "orders") or has_perm(user, "view_stats")):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    # Extra check: only users with view_stats can see stats data
    if not has_perm(user, "view_stats"):
        return JSONResponse({"error": "No stats permission"}, status_code=403)
    pool = await get_db_pool()
    d_from = date.fromisoformat(date_from) if date_from else date.today() - timedelta(days=30)
    d_to   = date.fromisoformat(date_to)   if date_to   else date.today()

    # Build source filter for orders
    source_filter = ""
    params = [d_from, d_to]
    if source or seller:
        sources = [s.strip() for s in source.split(",") if s.strip()] if source else []
        sellers = [s.strip() for s in seller.split(",") if s.strip()] if seller else []
        if sources and sellers:
            # Combined: (non-allegro sources) OR (allegro + specific sellers)
            non_allegro = [s for s in sources if s != "allegro"]
            or_parts = []
            if non_allegro:
                params.append(non_allegro)
                or_parts.append(f"source = ANY(${len(params)})")
            params.append(sellers)
            or_parts.append(f"(source = 'allegro' AND seller_account = ANY(${len(params)}))")
            source_filter += f" AND ({' OR '.join(or_parts)})"
        elif sources:
            params.append(sources)
            source_filter += f" AND source = ANY(${len(params)})"
        elif sellers:
            params.append(sellers)
            source_filter += f" AND (source = 'allegro' AND seller_account = ANY(${len(params)}))"

    async with pool.acquire() as conn:
        leads = await conn.fetchrow("""
            SELECT
                COUNT(*)                                              AS total,
                COUNT(*) FILTER (WHERE form_type = 'call')            AS calls,
                COUNT(*) FILTER (WHERE form_type IN ('form','chat'))  AS forms,
                COUNT(*) FILTER (WHERE status = 'new')                AS new_leads,
                COUNT(*) FILTER (WHERE status = 'processed')          AS processed
            FROM leads
            WHERE created_at::date >= $1 AND created_at::date <= $2
        """, d_from, d_to)

        orders = await conn.fetchrow(f"""
            SELECT
                COUNT(*) AS total,
                COALESCE(SUM(total), 0) AS revenue,
                COUNT(DISTINCT customer_email) FILTER (WHERE customer_email != '') AS unique_customers,
                COUNT(*) FILTER (WHERE source = 'dobraszklarnia') AS woo_count,
                COUNT(*) FILTER (WHERE source = 'oteko') AS oteko_count,
                COUNT(*) FILTER (WHERE source = 'cieplarnia') AS ciep_count,
                COUNT(*) FILTER (WHERE source = 'allegro') AS allegro_count
            FROM orders
            WHERE COALESCE(external_created, created_at)::date >= $1
              AND COALESCE(external_created, created_at)::date <= $2
              {source_filter}
        """, *params)

    return JSONResponse({
        "leads": _safe_dict(leads),
        "orders": _safe_dict(orders),
        "date_from": d_from.isoformat(),
        "date_to": d_to.isoformat(),
    })


# ============================================================
# === API: Статистика по дням (для графика) ===
# ============================================================


@app.get("/api/dashboard/stats")
async def api_dashboard_stats(
    request: Request,
    date_from: str = Query(None),
    date_to:   str = Query(None),
):
    """Dynamic dashboard stats: totals + breakdown by source + by seller for allegro."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "dashboard"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    pool = await get_db_pool()
    d_from = date.fromisoformat(date_from) if date_from else date.today() - timedelta(days=90)
    d_to   = date.fromisoformat(date_to)   if date_to   else date.today()

    show_prices = has_perm(user, "prices")

    async with pool.acquire() as conn:
        # Total orders + revenue for the period
        totals = await conn.fetchrow("""
            SELECT
                COUNT(*) AS total,
                COALESCE(SUM(total), 0) AS revenue,
                COUNT(*) FILTER (WHERE COALESCE(external_created, created_at)::date = CURRENT_DATE) AS today_orders
            FROM orders
            WHERE COALESCE(external_created, created_at)::date >= $1
              AND COALESCE(external_created, created_at)::date <= $2
        """, d_from, d_to)

        # Breakdown by source
        by_source = await conn.fetch("""
            SELECT source, COUNT(*) AS cnt, COALESCE(SUM(total), 0) AS revenue
            FROM orders
            WHERE COALESCE(external_created, created_at)::date >= $1
              AND COALESCE(external_created, created_at)::date <= $2
            GROUP BY source ORDER BY cnt DESC
        """, d_from, d_to)

        # Breakdown by Allegro seller accounts
        by_seller = await conn.fetch("""
            SELECT seller_account, COUNT(*) AS cnt, COALESCE(SUM(total), 0) AS revenue
            FROM orders
            WHERE source = 'allegro'
              AND COALESCE(external_created, created_at)::date >= $1
              AND COALESCE(external_created, created_at)::date <= $2
              AND seller_account IS NOT NULL AND seller_account != ''
            GROUP BY seller_account ORDER BY cnt DESC
        """, d_from, d_to)

        # Breakdown by traffic channel (utm_source/utm_medium) per store source
        by_channel = await conn.fetch("""
            SELECT
                source,
                COALESCE(NULLIF(utm_source, ''), 'direct / unknown') AS channel,
                COALESCE(NULLIF(utm_medium, ''), '') AS medium,
                COUNT(*) AS cnt,
                COALESCE(SUM(total), 0) AS revenue
            FROM orders
            WHERE COALESCE(external_created, created_at)::date >= $1
              AND COALESCE(external_created, created_at)::date <= $2
            GROUP BY source, channel, medium
            ORDER BY source, cnt DESC
        """, d_from, d_to)

        # Breakdown by campaign (for Google Ads detail)
        by_campaign = await conn.fetch("""
            SELECT
                source,
                COALESCE(NULLIF(utm_source, ''), '') AS utm_src,
                COALESCE(NULLIF(utm_campaign, ''), '(not set)') AS campaign,
                COUNT(*) AS cnt,
                COALESCE(SUM(total), 0) AS revenue
            FROM orders
            WHERE COALESCE(external_created, created_at)::date >= $1
              AND COALESCE(external_created, created_at)::date <= $2
              AND utm_source IS NOT NULL AND utm_source != ''
              AND utm_campaign IS NOT NULL AND utm_campaign != ''
            GROUP BY source, utm_src, campaign
            ORDER BY cnt DESC
            LIMIT 30
        """, d_from, d_to)

        # Device breakdown
        by_device = await conn.fetch("""
            SELECT
                source,
                COALESCE(NULLIF(device_type, ''), 'Unknown') AS device,
                COUNT(*) AS cnt,
                COALESCE(SUM(total), 0) AS revenue
            FROM orders
            WHERE COALESCE(external_created, created_at)::date >= $1
              AND COALESCE(external_created, created_at)::date <= $2
              AND device_type IS NOT NULL AND device_type != ''
            GROUP BY source, device
            ORDER BY source, cnt DESC
        """, d_from, d_to)

    result = {
        "total_orders": int(totals["total"]),
        "today_orders": int(totals["today_orders"]),
        "revenue": float(totals["revenue"]) if show_prices else None,
        "by_source": [
            {
                "source": r["source"],
                "cnt": int(r["cnt"]),
                "revenue": float(r["revenue"]) if show_prices else None,
            }
            for r in by_source
        ],
        "by_seller": [
            {
                "seller": r["seller_account"],
                "cnt": int(r["cnt"]),
                "revenue": float(r["revenue"]) if show_prices else None,
            }
            for r in by_seller
        ],
        "by_channel": [
            {
                "source": r["source"],
                "channel": r["channel"],
                "medium": r["medium"],
                "cnt": int(r["cnt"]),
                "revenue": float(r["revenue"]) if show_prices else None,
            }
            for r in by_channel
        ],
        "by_campaign": [
            {
                "source": r["source"],
                "utm_source": r["utm_src"],
                "campaign": r["campaign"],
                "cnt": int(r["cnt"]),
                "revenue": float(r["revenue"]) if show_prices else None,
            }
            for r in by_campaign
        ],
        "by_device": [
            {
                "source": r["source"],
                "device": r["device"],
                "cnt": int(r["cnt"]),
                "revenue": float(r["revenue"]) if show_prices else None,
            }
            for r in by_device
        ],
    }
    return JSONResponse(result)


@app.get("/api/orders/daily")
async def api_orders_daily(
    request: Request,
    date_from: str = Query(None),
    date_to:   str = Query(None),
    source:    str = Query(None),
):
    user = await get_current_user(request)
    if not user or not (has_perm(user, "dashboard") or has_perm(user, "orders")):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    pool = await get_db_pool()
    d_from = date.fromisoformat(date_from) if date_from else date.today() - timedelta(days=30)
    d_to   = date.fromisoformat(date_to)   if date_to   else date.today()

    conditions = ["COALESCE(external_created, created_at)::date >= $1", "COALESCE(external_created, created_at)::date <= $2"]
    params = [d_from, d_to]
    if source:
        params.append(source)
        conditions.append(f"source = ${len(params)}")

    where = " AND ".join(conditions)

    async with pool.acquire() as conn:
        rows = await conn.fetch(f"""
            SELECT
                COALESCE(external_created, created_at)::date AS day,
                COUNT(*) AS orders_count,
                COALESCE(SUM(total), 0) AS revenue
            FROM orders
            WHERE {where}
            GROUP BY day
            ORDER BY day
        """, *params)

    return JSONResponse({
        "days": [r["day"].isoformat() for r in rows],
        "orders_count": [int(r["orders_count"]) for r in rows],
        "revenue": [float(r["revenue"]) for r in rows],
    })


# ============================================================
# === API: Ручное добавление заказа ===
# ============================================================

@app.post("/api/orders/manual")
async def add_manual_order(request: Request):
    """Добавить заказ вручную. Не перезаписывается синком."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "add_orders"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "invalid json"})

    pool = await get_db_pool()

    # Generate unique external_id for manual orders
    import uuid
    ext_id = f"manual-{uuid.uuid4().hex[:8]}"
    source = body.get("source", "manual")

    # Parse date
    ext_created = None
    if body.get("date"):
        try:
            ext_created = datetime.fromisoformat(body["date"])
        except Exception:
            ext_created = datetime.now(TZ)
    else:
        ext_created = datetime.now(TZ)

    price = float(body.get("price", 0) or 0)
    ship = float(body.get("shipping_cost", 0) or 0)
    total = price + ship

    items = [{
        "name": body.get("product_name", ""),
        "qty": 1,
        "price": str(price),
        "sku": "",
    }]

    order_data = {
        "external_id": ext_id,
        "status": body.get("status", "processing"),
        "customer_name": body.get("customer_name", ""),
        "customer_email": body.get("customer_email", ""),
        "customer_phone": body.get("customer_phone", ""),
        "customer_city": body.get("customer_city", ""),
        "customer_zip": "",
        "customer_address": body.get("customer_address", ""),
        "total": total,
        "currency": "PLN",
        "items_count": 1,
        "items_json": json.dumps(items, ensure_ascii=False),
        "payment_method": body.get("payment_method", ""),
        "shipping_method": "",
        "note": body.get("comment", ""),
        "external_created": ext_created,
        "shipping_cost": ship,
        "customer_comment": body.get("comment", ""),
        "seller_account": body.get("seller_account", ""),
        "product_type": body.get("product_type", ""),
        "configuration": body.get("configuration", ""),
    }

    async with pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO orders (
                external_id, source, status, customer_name, customer_email,
                customer_phone, customer_city, customer_zip, customer_address,
                total, currency, items_count, items_json,
                payment_method, shipping_method, note, external_created,
                shipping_cost, seller_account, customer_comment,
                product_type, configuration, created_by_email
            ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18,$19,$20,$21,$22,$23)
        """,
            ext_id, source,
            order_data["status"], order_data["customer_name"], order_data["customer_email"],
            order_data["customer_phone"], order_data["customer_city"], order_data["customer_zip"],
            order_data["customer_address"], total, "PLN", 1, order_data["items_json"],
            order_data["payment_method"], "", order_data["note"], ext_created,
            ship, order_data["seller_account"], order_data["customer_comment"],
            order_data["product_type"], order_data["configuration"],
            user.get("email", ""),
        )

    return JSONResponse({"ok": True, "id": ext_id})


@app.put("/api/orders/manual/{order_id}")
async def edit_manual_order(order_id: int, request: Request):
    """Edit a manually created order. Creator or superadmin can edit."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "add_orders"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    body = await request.json()
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT source, external_id, created_by_email FROM orders WHERE id=$1", order_id)
        if not row or not (row["external_id"] or "").startswith("manual"):
            return JSONResponse({"ok": False, "error": "Only manual orders can be edited"}, status_code=400)
        # Only creator or superadmin can edit
        if not has_perm(user, "manage_users"):
            if (row["created_by_email"] or "") != (user.get("email", "") or ""):
                return JSONResponse({"ok": False, "error": "Możesz edytować tylko swoje zamówienia"}, status_code=403)

        price = float(body.get("price", 0) or 0)
        ship = float(body.get("shipping_cost", 0) or 0)
        total = price + ship
        items = [{"name": body.get("product_name", ""), "qty": 1, "price": str(price), "sku": ""}]

        ext_created = None
        if body.get("date"):
            try:
                ext_created = datetime.fromisoformat(body["date"])
            except Exception:
                ext_created = None

        await conn.execute("""
            UPDATE orders SET
                status=$1, customer_name=$2, customer_phone=$3, customer_city=$4,
                customer_address=$5, total=$6, items_json=$7, payment_method=$8,
                customer_comment=$9, shipping_cost=$10, seller_account=$11,
                product_type=$12, configuration=$13, customer_email=$14,
                external_created=COALESCE($15, external_created),
                source=$16, updated_at=NOW()
            WHERE id=$17
        """,
            body.get("status", "processing"),
            body.get("customer_name", ""),
            body.get("customer_phone", ""),
            body.get("customer_city", ""),
            body.get("customer_address", ""),
            total,
            json.dumps(items, ensure_ascii=False),
            body.get("payment_method", ""),
            body.get("comment", ""),
            ship,
            body.get("seller_account", ""),
            body.get("product_type", ""),
            body.get("configuration", ""),
            body.get("customer_email", ""),
            ext_created,
            body.get("source", "manual"),
            order_id,
        )
    return JSONResponse({"ok": True})


@app.delete("/api/orders/manual/{order_id}")
async def delete_manual_order(order_id: int, request: Request):
    """Delete a manually created order. Only superadmin (manage_users) can delete."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "manage_users"):
        return JSONResponse({"error": "Tylko superadmin może usuwać zamówienia"}, status_code=403)
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT external_id FROM orders WHERE id=$1", order_id)
        if not row or not (row["external_id"] or "").startswith("manual"):
            return JSONResponse({"ok": False, "error": "Only manual orders can be deleted"}, status_code=400)
        await conn.execute("DELETE FROM orders WHERE id=$1", order_id)
    return JSONResponse({"ok": True})


# ============================================================
# === API: Update order comment ===
# ============================================================

@app.post("/api/orders/update-comment")
async def update_order_comment(request: Request):
    """Update comment for any order."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "edit_comments"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    body = await request.json()
    order_id = body.get("id")
    comment = body.get("comment", "")
    if not order_id:
        return JSONResponse({"ok": False, "error": "no id"})
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE orders SET customer_comment=$1 WHERE id=$2",
            comment, int(order_id),
        )
    return JSONResponse({"ok": True})


# ============================================================
# === API: Toggle exported status ===
# ============================================================

@app.post("/api/orders/toggle-exported")
async def toggle_exported(request: Request):
    """Toggle exported flag for an order."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "mark_delivery"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    body = await request.json()
    order_id = body.get("id")
    exported = body.get("exported", False)
    if not order_id:
        return JSONResponse({"ok": False, "error": "no id"})
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE orders SET exported=$1, exported_at=$2 WHERE id=$3",
            exported,
            datetime.now() if exported else None,
            int(order_id),
        )
    return JSONResponse({"ok": True})


@app.post("/api/orders/mark-exported")
async def mark_exported(request: Request):
    """Mark multiple orders as exported (called after XLSX export)."""
    body = await request.json()
    ids = body.get("ids", [])
    if not ids:
        return JSONResponse({"ok": False, "error": "no ids"})
    int_ids = [int(x) for x in ids if str(x).isdigit()]
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE orders SET exported=TRUE, exported_at=$1 WHERE id = ANY($2)",
            datetime.now(), int_ids,
        )
    return JSONResponse({"ok": True, "marked": len(int_ids)})


# ── Export history & undo ──

@app.get("/api/export-log")
async def api_export_log(request: Request, limit: int = Query(20)):
    """Get recent export log entries for undo capability."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "export"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, user_email, orders_count, created_at, undone, undone_at FROM export_log ORDER BY id DESC LIMIT $1",
            limit
        )
    log = []
    for r in rows:
        log.append({
            "id": r["id"],
            "user_email": r["user_email"],
            "orders_count": r["orders_count"],
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            "undone": r["undone"],
            "undone_at": r["undone_at"].isoformat() if r.get("undone_at") else None,
        })
    return JSONResponse({"log": log})


@app.post("/api/export-log/{log_id}/undo")
async def api_undo_export(log_id: int, request: Request):
    """Undo an export — reset exported flag for all orders in that export batch."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "manage_users"):
        return JSONResponse({"error": "Only admin can undo exports"}, status_code=403)
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM export_log WHERE id=$1", log_id)
        if not row:
            return JSONResponse({"error": "Export log not found"}, status_code=404)
        if row["undone"]:
            return JSONResponse({"error": "Already undone"}, status_code=400)
        order_ids = row["order_ids"]
        if order_ids:
            await conn.execute(
                "UPDATE orders SET exported=FALSE, exported_at=NULL WHERE id = ANY($1)",
                order_ids,
            )
        await conn.execute(
            "UPDATE export_log SET undone=TRUE, undone_at=$1 WHERE id=$2",
            datetime.now(), log_id,
        )
    return JSONResponse({"ok": True, "reverted": len(order_ids) if order_ids else 0})


@app.post("/api/orders/reset-all-exported")
async def api_reset_all_exported(request: Request):
    """Emergency reset: clear ALL exported flags. Admin only."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "manage_users"):
        return JSONResponse({"error": "Only admin can do this"}, status_code=403)
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        result = await conn.execute(
            "UPDATE orders SET exported=FALSE, exported_at=NULL WHERE exported=TRUE"
        )
        count = int(result.split()[-1]) if result else 0
    return JSONResponse({"ok": True, "reset_count": count})


@app.get("/api/orders/exported-timestamps")
async def api_exported_timestamps(request: Request):
    """Show mass-export timestamps to help identify the bulk export to undo."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "manage_users"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        # Group by exported_at rounded to the second — mass exports will have hundreds at the same second
        rows = await conn.fetch("""
            SELECT date_trunc('second', exported_at) AS ts,
                   COUNT(*) AS cnt
            FROM orders
            WHERE exported = TRUE AND exported_at IS NOT NULL
            GROUP BY date_trunc('second', exported_at)
            ORDER BY cnt DESC
            LIMIT 20
        """)
    result = []
    for r in rows:
        result.append({
            "timestamp": r["ts"].isoformat() if r["ts"] else None,
            "count": r["cnt"],
        })
    return JSONResponse({"timestamps": result})


@app.post("/api/orders/revert-bulk-export")
async def api_revert_bulk_export(request: Request):
    """Revert a mass export by timestamp — only reset orders exported at that exact second.
    Orders exported manually at other times are preserved."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "manage_users"):
        return JSONResponse({"error": "Only admin can do this"}, status_code=403)
    body = await request.json()
    ts = body.get("timestamp")
    if not ts:
        return JSONResponse({"error": "timestamp required"}, status_code=400)
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        target_ts = datetime.fromisoformat(ts)
        # Reset only orders exported at that exact second (±1 second tolerance)
        result = await conn.execute(
            """UPDATE orders SET exported=FALSE, exported_at=NULL
               WHERE exported=TRUE
                 AND exported_at >= $1
                 AND exported_at < $2""",
            target_ts, target_ts + timedelta(seconds=2),
        )
        count = int(result.split()[-1]) if result else 0
    return JSONResponse({"ok": True, "reverted": count, "timestamp": ts})


@app.post("/api/orders/restore-exported")
async def api_restore_exported(request: Request):
    """Restore exported flags for specific order IDs. Admin only."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "manage_users"):
        return JSONResponse({"error": "Only admin can do this"}, status_code=403)
    body = await request.json()
    ids = body.get("ids", [])
    if not ids:
        return JSONResponse({"error": "ids required"}, status_code=400)
    int_ids = [int(x) for x in ids if str(x).isdigit()]
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE orders SET exported=TRUE, exported_at=$1 WHERE id = ANY($2)",
            datetime.now(), int_ids,
        )
    return JSONResponse({"ok": True, "restored": len(int_ids)})


# One-time recovery: IDs from legitimate exports on 2026-06-01 (before mass export)
_RECOVERY_IDS = [4223,4225,4226,4227,4228,4229,4231,4232,4234,4237,4239,4242,4243,4246,4250,4256,4257,4265,4271,4273,4274,4275,4279,4284,4285,4286,4287,4290,4293,4295,4296,4297,4298,4302,4303,4306,4307,4308,9686,9690,9693,9697,9700,9706,9710,9719,9721,9746,9747,9756,9762,9767,9769,9774,9775,9777,9782,9783,9792,9800,9815,9831,9832,9843,9848,9849,9850,9872,9873,9876,9878,9881,9885,9887,9892,9895,9905,9906,9908,9909,9910]


@app.get("/api/orders/recover-june1")
async def api_recover_june1(request: Request):
    """One-time recovery: restore 81 orders from legitimate exports on June 1."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "manage_users"):
        return JSONResponse({"error": "Only admin can do this"}, status_code=403)
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE orders SET exported=TRUE, exported_at=$1 WHERE id = ANY($2)",
            datetime.now(), _RECOVERY_IDS,
        )
    return JSONResponse({"ok": True, "restored": len(_RECOVERY_IDS), "ids": _RECOVERY_IDS})


@app.post("/api/orders/toggle-priority")
async def toggle_priority(request: Request):
    """Toggle priority flag for an order."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "orders"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    body = await request.json()
    order_id = body.get("id")
    is_priority = body.get("is_priority", False)
    if not order_id:
        return JSONResponse({"ok": False, "error": "no id"})
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE orders SET is_priority=$1 WHERE id=$2",
            is_priority, int(order_id),
        )
    return JSONResponse({"ok": True})


# ============================================================
# === API: Заказы (для таблицы + экспорт) ===
# ============================================================

@app.get("/api/orders")
async def api_orders(
    request: Request,
    limit:  int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    source: str = Query(None),
    seller: str = Query(None),
    status: str = Query(None),
    search: str = Query(None),
    date_from: str = Query(None),
    date_to:   str = Query(None),
    exported: str = Query(None),
    priority: str = Query(None),
    product_type: str = Query(None),
):
    user = await get_current_user(request)
    if not user:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    if not has_perm(user, "orders"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    pool = await get_db_pool()
    conditions = ["1=1"]
    params = []

    # Filter by allowed sources for this user
    # allowed_sources can contain: "dobraszklarnia,allegro,allegro:mantrade,allegro:drogatrade"
    # Plain values (e.g. "allegro") filter by source column
    # Colon values (e.g. "allegro:mantrade") filter by source + seller_account
    user_sources = user.get("allowed_sources", "all") if user else "all"
    if user_sources and user_sources != "all":
        src_list = [s.strip() for s in user_sources.split(",") if s.strip()]
        if src_list:
            plain_sources = [s for s in src_list if ":" not in s]
            seller_pairs = [s.split(":", 1) for s in src_list if ":" in s]
            src_conds = []
            if plain_sources:
                params.append(plain_sources)
                src_conds.append(f"source = ANY(${len(params)})")
            for src, slr in seller_pairs:
                params.append(src)
                params.append(slr)
                src_conds.append(f"(source = ${len(params)-1} AND seller_account = ${len(params)})")
            if src_conds:
                conditions.append(f"({' OR '.join(src_conds)})")

    if source or seller:
        source_list = [s.strip() for s in source.split(",") if s.strip()] if source else []
        seller_list = [s.strip() for s in seller.split(",") if s.strip()] if seller else []
        if source_list and seller_list:
            # Combined filter: (non-allegro sources) OR (allegro + specific sellers)
            non_allegro = [s for s in source_list if s != "allegro"]
            or_parts = []
            if non_allegro:
                params.append(non_allegro)
                or_parts.append(f"source = ANY(${len(params)})")
            params.append(seller_list)
            or_parts.append(f"(source = 'allegro' AND seller_account = ANY(${len(params)}))")
            conditions.append(f"({' OR '.join(or_parts)})")
        elif source_list:
            if len(source_list) == 1:
                params.append(source_list[0])
                conditions.append(f"source = ${len(params)}")
            else:
                params.append(source_list)
                conditions.append(f"source = ANY(${len(params)})")
        elif seller_list:
            params.append(seller_list)
            conditions.append(f"(source = 'allegro' AND seller_account = ANY(${len(params)}))")
    if status:
        params.append(status)
        conditions.append(f"status = ${len(params)}")
    if search:
        params.append(f"%{search}%")
        conditions.append(f"(customer_name ILIKE ${len(params)} OR customer_phone ILIKE ${len(params)} OR customer_email ILIKE ${len(params)})")
    if date_from:
        params.append(date.fromisoformat(date_from))
        conditions.append(f"COALESCE(external_created, created_at)::date >= ${len(params)}")
    if date_to:
        params.append(date.fromisoformat(date_to))
        conditions.append(f"COALESCE(external_created, created_at)::date <= ${len(params)}")
    if exported == "yes":
        conditions.append("exported = TRUE")
    elif exported == "no":
        conditions.append("(exported IS NULL OR exported = FALSE)")
    if priority == "yes":
        conditions.append("is_priority = TRUE")
    if product_type:
        params.append(product_type)
        conditions.append(f"product_type = ${len(params)}")

    where = " AND ".join(conditions)

    params.extend([limit, offset])
    q = f"""
        SELECT * FROM orders
        WHERE {where}
        ORDER BY is_priority DESC NULLS LAST, COALESCE(external_created, created_at) DESC
        LIMIT ${len(params)-1} OFFSET ${len(params)}
    """

    async with pool.acquire() as conn:
        rows = await conn.fetch(q, *params)
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM orders WHERE {where}",
            *params[:-2]
        )

    orders = [_safe_dict(r) for r in rows]

    # Hide prices if no "prices" permission
    if not has_perm(user, "prices"):
        for o in orders:
            o["total"] = None
            o["shipping_cost"] = None

    # Hide phones if no "phones" permission
    if not has_perm(user, "phones"):
        for o in orders:
            o["customer_phone"] = None
            o["customer_email"] = None

    return JSONResponse({
        "total": total,
        "orders": orders,
        "permissions": user.get("permissions", ""),
    })


# ============================================================
# === API: Диагностика NIP ===
# ============================================================

@app.get("/api/debug/utm")
async def debug_utm(request: Request):
    """Диагностика: показать raw meta_data из 5 последних WooCommerce заказов (ищем UTM/attribution)."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "manage_users"):
        return JSONResponse({"error": "superadmin only"}, status_code=403)

    # Fetch 5 latest WooCommerce orders raw from API
    if not WOO_KEY or not WOO_SECRET:
        return JSONResponse({"error": "WooCommerce not configured"})

    results = []
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                f"{WOO_URL}/wp-json/wc/v3/orders",
                params={"per_page": 5, "orderby": "date", "order": "desc"},
                auth=aiohttp.BasicAuth(WOO_KEY, WOO_SECRET),
                ssl=False, timeout=aiohttp.ClientTimeout(total=30)
            ) as resp:
                if resp.status != 200:
                    return JSONResponse({"error": f"WooCommerce API returned {resp.status}"})
                orders = await resp.json()

                for woo in orders:
                    meta = woo.get("meta_data", [])
                    # Filter to show only attribution-related meta keys
                    attribution_meta = [
                        {"key": m.get("key"), "value": m.get("value")}
                        for m in meta
                        if any(kw in (m.get("key") or "").lower() for kw in
                               ("utm", "attribution", "source", "medium", "campaign", "referrer", "gclid", "fbclid", "session"))
                    ]
                    results.append({
                        "order_id": woo.get("id"),
                        "date": woo.get("date_created"),
                        "attribution_meta": attribution_meta,
                        "all_meta_keys": [m.get("key") for m in meta],
                    })
    except Exception as e:
        return JSONResponse({"error": str(e)})

    # Also show current DB UTM stats
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        utm_stats = await conn.fetch("""
            SELECT source,
                   COUNT(*) FILTER (WHERE utm_source IS NOT NULL AND utm_source != '') AS has_utm,
                   COUNT(*) AS total
            FROM orders GROUP BY source ORDER BY source
        """)

    return JSONResponse({
        "woo_raw_samples": results,
        "db_utm_stats": [{"source": r["source"], "has_utm": r["has_utm"], "total": r["total"]} for r in utm_stats],
    })


@app.get("/api/debug/nip")
async def debug_nip(request: Request):
    """Диагностика: показать заказы с NIP в БД."""
    user = await get_current_user(request)
    if not user or not has_perm(user, "manage_users"):
        return JSONResponse({"error": "superadmin only"}, status_code=403)
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, source, external_id, customer_name, invoice_nip FROM orders WHERE invoice_nip IS NOT NULL AND invoice_nip != '' LIMIT 50"
        )
        total = await conn.fetchval(
            "SELECT COUNT(*) FROM orders WHERE invoice_nip IS NOT NULL AND invoice_nip != ''"
        )
        by_source = await conn.fetch(
            "SELECT source, COUNT(*) as cnt FROM orders WHERE invoice_nip IS NOT NULL AND invoice_nip != '' GROUP BY source"
        )
    return JSONResponse({
        "total_with_nip": total,
        "by_source": {r["source"]: r["cnt"] for r in by_source},
        "samples": [_safe_dict(r) for r in rows]
    })


# ============================================================
# === API: Экспорт заказов в CSV ===
# ============================================================

@app.get("/api/orders/export")
async def export_orders(
    request: Request,
    source: str = Query(None),
    seller: str = Query(None),
    status: str = Query(None),
    exported: str = Query(None),
    date_from: str = Query(None),
    date_to:   str = Query(None),
    ids: str = Query(None),
    fmt: str = Query("csv"),
):
    user = await get_current_user(request)
    if not user or not has_perm(user, "export"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    pool = await get_db_pool()
    conditions = ["1=1"]
    params = []

    # Export selected IDs (checkboxes)
    if ids:
        id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        if id_list:
            params.append(id_list)
            conditions.append(f"id = ANY(${len(params)})")

    if source:
        params.append(source)
        conditions.append(f"source = ${len(params)}")
    if seller:
        params.append(seller)
        conditions.append(f"seller_account = ${len(params)}")
    if status:
        params.append(status)
        conditions.append(f"status = ${len(params)}")
    if exported == "yes":
        conditions.append("exported = TRUE")
    elif exported == "no":
        conditions.append("(exported IS NULL OR exported = FALSE)")
    if date_from:
        params.append(date.fromisoformat(date_from))
        conditions.append(f"COALESCE(external_created, created_at)::date >= ${len(params)}")
    if date_to:
        params.append(date.fromisoformat(date_to))
        conditions.append(f"COALESCE(external_created, created_at)::date <= ${len(params)}")

    where = " AND ".join(conditions)

    async with pool.acquire() as conn:
        rows = await conn.fetch(
            f"SELECT * FROM orders WHERE {where} ORDER BY COALESCE(external_created, created_at) DESC",
            *params
        )
        # Auto-mark exported orders + log for undo
        if rows:
            exported_ids = [r["id"] for r in rows]
            await conn.execute(
                "UPDATE orders SET exported=TRUE, exported_at=$1 WHERE id = ANY($2)",
                datetime.now(), exported_ids,
            )
            # Log this export for undo capability
            await conn.execute(
                "INSERT INTO export_log (user_email, order_ids, orders_count) VALUES ($1, $2, $3)",
                user.get("email", ""), exported_ids, len(exported_ids),
            )

    # --- XLSX export (openpyxl) — per-item breakdown ---
    import re as _re_export
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Zamówienia"

    # Header row — per-item breakdown format
    headers = [
        "Data", "Nr zamówienia", "Źródło", "Konto",
        "Klient", "Telefon", "Adres",
        "Produkt", "SKU", "Ilość", "Cena netto szt.", "Cena brutto szt.", "Wartość brutto",
        "Dostawa", "Razem zamówienie",
        "Typ towaru", "Grubość (mm)", "Konfiguracja",
        "Status", "Płatność", "NIP", "Komentarz"
    ]
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
    )
    subtotal_fill = PatternFill("solid", fgColor="F3F4F6")
    subtotal_font = Font(name="Arial", bold=True, size=10)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Helper: extract polycarbonate thickness from product name/sku
    def _extract_thickness(name: str, sku: str) -> str:
        combined = f"{name} {sku}".lower()
        m = _re_export.search(r'(\d+)\s*mm', combined)
        if m:
            return m.group(1)
        return ""

    # Data rows — each item gets its own row
    data_font = Font(name="Arial", size=10)
    date_fmt = "DD.MM.YYYY HH:MM"
    current_row = 2

    for r in rows:
        items = []
        try:
            items = json.loads(r["items_json"] or "[]")
        except Exception:
            pass
        if not items:
            items = [{"name": "", "full_name": "", "qty": 1, "price": "0", "sku": ""}]

        total_val = float(r["total"] or 0)
        ship_cost = float(r["shipping_cost"] or 0) if r.get("shipping_cost") else 0.0
        is_allegro = (r["source"] or "").lower() == "allegro"

        # Address
        addr_parts = []
        if r["customer_city"]:
            addr_parts.append(r["customer_city"])
        if r["customer_address"]:
            addr_parts.append(r["customer_address"])
        if r["customer_zip"]:
            addr_parts.append(r["customer_zip"])
        addr_str = ", ".join(addr_parts)

        comment = r.get("customer_comment") or r.get("note") or ""
        status_pl = _status_pl(r["status"])
        order_date = r["external_created"] if r["external_created"] else r["created_at"]
        order_id = r.get("external_id") or str(r.get("id", ""))
        source_label = r["source"] or ""
        if r.get("seller_account"):
            source_label = f"{source_label} ({r['seller_account']})"

        for item_idx, it in enumerate(items):
            item_name = it.get("full_name") or it.get("name") or ""
            item_sku = it.get("sku") or ""
            item_qty = int(it.get("qty", 1))
            item_price_raw = float(it.get("price") or 0)

            # Calculate netto/brutto per item
            if is_allegro:
                # Allegro prices are already brutto
                brutto_per_unit = item_price_raw
                netto_per_unit = round(item_price_raw / 1.23, 2)
            else:
                # WooCommerce/PrestaShop prices are netto
                netto_per_unit = item_price_raw
                brutto_per_unit = round(item_price_raw * 1.23, 2)

            brutto_total = round(brutto_per_unit * item_qty, 2)
            thickness = _extract_thickness(item_name, item_sku)

            row_data = [
                order_date if item_idx == 0 else "",                       # Data (only first item)
                order_id if item_idx == 0 else "",                         # Nr zamówienia
                (r["source"] or "") if item_idx == 0 else "",              # Źródło
                (r.get("seller_account") or "") if item_idx == 0 else "",  # Konto
                (r.get("customer_name") or "") if item_idx == 0 else "",   # Klient
                (r["customer_phone"] or "") if item_idx == 0 else "",      # Telefon
                addr_str if item_idx == 0 else "",                         # Adres
                item_name,                                                 # Produkt
                item_sku,                                                  # SKU
                item_qty,                                                  # Ilość
                netto_per_unit,                                            # Cena netto szt.
                brutto_per_unit,                                           # Cena brutto szt.
                brutto_total,                                              # Wartość brutto
                ship_cost if item_idx == 0 else "",                        # Dostawa (only first)
                total_val if item_idx == 0 else "",                        # Razem zamówienie (only first)
                (r.get("product_type") or "") if item_idx == 0 else "",    # Typ towaru
                thickness,                                                 # Grubość (mm)
                (r.get("configuration") or "") if item_idx == 0 else "",   # Konfiguracja
                status_pl if item_idx == 0 else "",                        # Status
                (r.get("payment_method") or "") if item_idx == 0 else "",  # Płatność
                (r.get("invoice_nip") or "") if item_idx == 0 else "",     # NIP
                comment if item_idx == 0 else "",                          # Komentarz
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border

            # Format date cell
            date_cell = ws.cell(row=current_row, column=1)
            if date_cell.value and hasattr(date_cell.value, 'strftime'):
                date_cell.number_format = date_fmt

            # Format money cells
            for money_col in [11, 12, 13, 14, 15]:
                c = ws.cell(row=current_row, column=money_col)
                if c.value != "":
                    c.number_format = '#,##0.00'

            current_row += 1

    # Column widths
    col_widths = [18, 14, 14, 14, 20, 16, 35, 45, 14, 8, 14, 14, 14, 12, 14, 14, 10, 16, 18, 20, 16, 30]
    for i, w in enumerate(col_widths, 1):
        if i <= len(headers):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    last_col_letter = ws.cell(row=1, column=len(headers)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{current_row - 1}"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"orders_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================
# === API: Клиенты ===
# ============================================================

@app.get("/api/customers")
async def api_customers(
    limit:  int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    search: str = Query(None),
):
    pool = await get_db_pool()
    conditions = ["1=1"]
    params = []

    if search:
        params.append(f"%{search}%")
        conditions.append(f"(name ILIKE ${len(params)} OR phone ILIKE ${len(params)} OR email ILIKE ${len(params)})")

    where = " AND ".join(conditions)
    params.extend([limit, offset])

    async with pool.acquire() as conn:
        rows = await conn.fetch(
            f"SELECT * FROM customers WHERE {where} ORDER BY last_order DESC NULLS LAST LIMIT ${len(params)-1} OFFSET ${len(params)}",
            *params
        )
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM customers WHERE {where}",
            *params[:-2]
        )

    return JSONResponse({"total": total, "customers": [_safe_dict(r) for r in rows]})


# ============================================================
# === API: Лиды (оставляем из оригинала) ===
# ============================================================

@app.get("/api/leads")
async def api_leads(
    limit:  int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    status: str = Query(None),
):
    pool = await get_db_pool()
    conditions = ["1=1"]
    params = []

    if status:
        params.append(status)
        conditions.append(f"status = ${len(params)}")

    params.extend([limit, offset])
    q = f"""
        SELECT * FROM leads
        WHERE {' AND '.join(conditions)}
        ORDER BY created_at DESC
        LIMIT ${len(params)-1} OFFSET ${len(params)}
    """

    async with pool.acquire() as conn:
        rows = await conn.fetch(q, *params)
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM leads WHERE {' AND '.join(conditions)}",
            *params[:-2]
        )

    return JSONResponse({
        "total": total,
        "leads": [dict(r) | {"created_at": r["created_at"].isoformat()} for r in rows],
    })


# ============================================================
# === Webhooks ===
# ============================================================

@app.api_route("/api/webhook/lead", methods=["GET", "POST"])
async def webhook_lead(request: Request):
    data = {}
    data.update(dict(request.query_params))
    raw = await request.body()
    if raw:
        ct = request.headers.get("content-type", "")
        if "json" in ct:
            data.update(json.loads(raw))
        elif "form" in ct or "urlencoded" in ct:
            from urllib.parse import parse_qs
            for k, v in parse_qs(raw.decode()).items():
                data[k] = v[0] if len(v) == 1 else v

    name  = data.get("name", "")
    phone = data.get("phone", "")
    email = data.get("email", "")

    utm_source   = data.get("utm_source", "")
    utm_medium   = data.get("utm_medium", "")
    utm_campaign = data.get("utm_campaign", "")
    utm_term     = data.get("utm_term", "")
    gclid        = data.get("gclid", "")
    landing      = data.get("landing_page") or data.get("referer") or ""

    if not phone and not email:
        return JSONResponse({"ok": False, "error": "phone or email required"}, status_code=400)

    pool = await get_db_pool()
    async with pool.acquire() as conn:
        lead_id = await conn.fetchval("""
            INSERT INTO leads (name, phone, email, utm_source, utm_medium,
                               utm_campaign, utm_term, gclid, landing_page)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
            RETURNING id
        """, name, phone, email, utm_source, utm_medium,
             utm_campaign, utm_term, gclid, landing)

    print(f"[webhook] new lead #{lead_id}: {phone or email}")
    return JSONResponse({"ok": True, "lead_id": lead_id})


@app.api_route("/api/webhook/call", methods=["GET", "POST"])
async def webhook_call(request: Request):
    raw = await request.body()
    data = dict(request.query_params)
    if raw:
        ct = request.headers.get("content-type", "")
        if "json" in ct:
            data.update(json.loads(raw))
        elif "form" in ct or "urlencoded" in ct:
            from urllib.parse import parse_qs
            for k, v in parse_qs(raw.decode()).items():
                data[k] = v[0]

    phone = data.get("phone", "")
    lead_id = None
    if phone:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            lead_id = await conn.fetchval(
                "SELECT id FROM leads WHERE phone = $1 ORDER BY created_at DESC LIMIT 1", phone
            )
            if lead_id:
                await conn.execute("""
                    UPDATE leads SET form_type = 'call', device = 'phone', updated_at = NOW()
                    WHERE id = $1
                """, lead_id)

    return JSONResponse({"ok": True, "matched_lead": lead_id})


@app.api_route("/api/webhook/woo", methods=["POST"])
async def webhook_woo(request: Request):
    """Webhook от WooCommerce — мгновенный приём нового заказа."""
    raw = await request.body()
    if not raw:
        return JSONResponse({"ok": False, "error": "empty body"}, status_code=400)

    # Верификация подписи (если задан WOO_WEBHOOK_SECRET)
    if WOO_WEBHOOK_SECRET:
        signature = request.headers.get("X-WC-Webhook-Signature", "")
        expected = b64encode(
            hmac.new(WOO_WEBHOOK_SECRET.encode(), raw, hashlib.sha256).digest()
        ).decode()
        if not hmac.compare_digest(signature, expected):
            print("[webhook-woo] invalid signature!")
            return JSONResponse({"ok": False, "error": "invalid signature"}, status_code=401)

    data = json.loads(raw)

    # WooCommerce шлёт ping при создании webhook — игнорируем
    if not data.get("id"):
        return JSONResponse({"ok": True, "note": "ping acknowledged"})

    source = "dobraszklarnia"
    parsed = _parse_woo_order(data)

    pool = await get_db_pool()
    async with pool.acquire() as conn:
        result = await _save_order(conn, source, parsed)
        if result == "new":
            await _upsert_customer(
                conn, source,
                parsed["customer_name"], parsed["customer_email"],
                parsed["customer_phone"], parsed["customer_city"],
                parsed["total"], parsed["external_created"]
            )

    print(f"[webhook-woo] order #{parsed['external_id']} — {result}")
    return JSONResponse({"ok": True, "order_id": parsed["external_id"], "result": result})


# ============================================================
# === Ручной синк + лог ===
# ============================================================

@app.get("/admin/sync")
async def manual_sync(request: Request, source: str = Query(None), full: bool = Query(False)):
    """Ручной запуск синхронизации. /admin/sync?source=allegro&full=true"""
    user = await get_current_user(request)
    if not user or not has_perm(user, "sync"):
        return JSONResponse({"error": "Access denied"}, status_code=403)
    if source == "dobraszklarnia":
        asyncio.create_task(sync_woocommerce(full=full))
        return JSONResponse({"ok": True, "message": f"WooCommerce sync started (full={full})"})
    elif source == "oteko":
        asyncio.create_task(sync_presta_oteko(full=full))
        return JSONResponse({"ok": True, "message": f"Oteko sync started (full={full})"})
    elif source == "cieplarnia":
        asyncio.create_task(sync_presta_ciep(full=full))
        return JSONResponse({"ok": True, "message": f"Cieplarnia sync started (full={full})"})

    elif source == "allegro":
        # Sync all Allegro accounts, or a specific one via ?seller=name
        seller = request.query_params.get("seller", "")
        if seller:
            acc = next((a for a in ALLEGRO_ACCOUNTS if a["name"] == seller), None)
            if acc:
                asyncio.create_task(_sync_allegro_account(acc, full=full))
                return JSONResponse({"ok": True, "message": f"Allegro:{seller} sync started (full={full})"})
            else:
                return JSONResponse({"ok": False, "error": f"Allegro account '{seller}' not found"}, status_code=404)
        asyncio.create_task(sync_allegro(full=full))
        names = [a["name"] for a in ALLEGRO_ACCOUNTS]
        return JSONResponse({"ok": True, "message": f"Allegro sync started for {names} (full={full})"})
    else:
        # Синк всех
        asyncio.create_task(sync_woocommerce(full=full))
        asyncio.create_task(sync_presta_oteko(full=full))
        asyncio.create_task(sync_presta_ciep(full=full))
        asyncio.create_task(sync_allegro(full=full))
        names = [a["name"] for a in ALLEGRO_ACCOUNTS]
        return JSONResponse({"ok": True, "message": f"All syncs started (full={full}), allegro accounts: {names}"})


@app.get("/api/sync-log")
async def api_sync_log(limit: int = Query(20)):
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT * FROM sync_log ORDER BY started_at DESC LIMIT $1", limit
        )

    return JSONResponse({"log": [_safe_dict(r) for r in rows]})


# ============================================================
# === Утилита: debug — raw WooCommerce order data ===
# ============================================================

@app.get("/admin/woo-debug")
async def woo_debug(order_id: str = Query(None)):
    """Show raw WooCommerce line_items with meta_data for debugging."""
    if not WOO_KEY or not WOO_SECRET:
        return JSONResponse({"error": "no woo keys"})
    async with aiohttp.ClientSession() as session:
        if order_id:
            url = f"{WOO_URL}/wp-json/wc/v3/orders/{order_id}"
        else:
            url = f"{WOO_URL}/wp-json/wc/v3/orders?per_page=3&orderby=date&order=desc"
        auth = aiohttp.BasicAuth(WOO_KEY, WOO_SECRET)
        async with session.get(url, auth=auth, ssl=False) as resp:
            data = await resp.json(content_type=None)
    # Extract only line_items for clarity
    if isinstance(data, list):
        result = []
        for o in data:
            result.append({
                "id": o.get("id"),
                "line_items": [{
                    "name": li.get("name"),
                    "sku": li.get("sku"),
                    "meta_data": li.get("meta_data", []),
                    "variation_id": li.get("variation_id"),
                    "product_id": li.get("product_id"),
                } for li in o.get("line_items", [])]
            })
        return JSONResponse(result)
    else:
        return JSONResponse({
            "id": data.get("id"),
            "line_items": [{
                "name": li.get("name"),
                "sku": li.get("sku"),
                "meta_data": li.get("meta_data", []),
                "variation_id": li.get("variation_id"),
                "product_id": li.get("product_id"),
            } for li in data.get("line_items", [])]
        })


# ============================================================
# === Утилита: массовое обновление seller_account ===
# ============================================================

@app.get("/admin/fix-seller-account")
async def fix_seller_account():
    """Fix empty seller_account → 'mantrade' for existing Allegro orders."""
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "UPDATE orders SET seller_account='mantrade' WHERE source='allegro' AND (seller_account IS NULL OR seller_account='')"
        )
    return JSONResponse({"ok": True, "updated": str(res)})


# ============================================================
# === Утилита: вытянуть статусы PrestaShop ===
# ============================================================

@app.get("/admin/presta-states")
async def presta_states():
    """Fetch order_states from both PrestaShop stores."""
    results = {}
    stores = [
        ("oteko", "https://oteko.pl", os.getenv("PRESTA_OTEKO_KEY", "")),
        ("cieplarnia", "https://cieplarnia.pl", os.getenv("PRESTA_CIEP_KEY", "")),
    ]
    async with aiohttp.ClientSession() as session:
        for name, base_url, key in stores:
            if not key:
                results[name] = "no key"
                continue
            try:
                auth = b64encode(f"{key}:".encode()).decode()
                url = f"{base_url}/api/order_states?output_format=JSON&display=full"
                async with session.get(url, headers={"Authorization": f"Basic {auth}"}, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                    if resp.status == 200:
                        data = await resp.json(content_type=None)
                        states_raw = data.get("order_states", [])
                        mapped = {}
                        for st in states_raw:
                            sid = str(st.get("id", ""))
                            # name может быть список (мультиязычность) или строка
                            nm = st.get("name", "")
                            if isinstance(nm, list):
                                nm = nm[0].get("value", "") if nm else ""
                            elif isinstance(nm, dict):
                                nm = nm.get("value", "") or nm.get("language", "")
                            mapped[sid] = nm
                        results[name] = mapped
                    else:
                        results[name] = f"HTTP {resp.status}"
            except Exception as e:
                results[name] = str(e)
    return JSONResponse(results)


# ============================================================
# === Миграции / Healthcheck ===
# ============================================================

@app.get("/admin/migrate")
async def run_migration(request: Request):
    user = await get_current_user(request)
    if not user or not has_perm(user, "manage_users"):
        return JSONResponse({"error": "Admin access required"}, status_code=403)
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        migrations = [
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS status VARCHAR(50) DEFAULT 'new'",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS campaign_id VARCHAR(100)",
            "ALTER TABLE leads ADD COLUMN IF NOT EXISTS ad_id VARCHAR(100)",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS shipping_method VARCHAR(200)",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS customer_zip VARCHAR(20)",
            "ALTER TABLE customers ADD COLUMN IF NOT EXISTS source VARCHAR(50)",
            # v2 — new fields for detailed order view
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS product_type VARCHAR(100) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS configuration TEXT DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS shipping_cost NUMERIC(10,2) DEFAULT 0",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS seller_account VARCHAR(200) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS customer_comment TEXT DEFAULT ''",
            # v3 — exported flag (order picked up for delivery)
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS exported BOOLEAN DEFAULT FALSE",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS exported_at TIMESTAMP",
            # v3.1 — priority flag
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS is_priority BOOLEAN DEFAULT FALSE",
            # v3.2 — invoice NIP + created_by
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS invoice_nip VARCHAR(50) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS created_by_email VARCHAR(200) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS utm_source VARCHAR(200) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS utm_medium VARCHAR(200) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS utm_campaign VARCHAR(500) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS utm_content VARCHAR(500) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS utm_term VARCHAR(500) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS traffic_source_type VARCHAR(50) DEFAULT ''",
            "ALTER TABLE orders ADD COLUMN IF NOT EXISTS device_type VARCHAR(50) DEFAULT ''",
            # v4 — user source access
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS allowed_sources TEXT DEFAULT 'all'",
            # v5 — permission-based access
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS permissions TEXT DEFAULT ''",
        ]
        results = []
        for m in migrations:
            try:
                await conn.execute(m)
                results.append(f"OK: {m[:80]}")
            except Exception as e:
                results.append(f"ERR: {m[:80]} — {e}")

    return JSONResponse({"migrations": results})


@app.get("/health")
async def health():
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.fetchval("SELECT 1")
        return JSONResponse({"status": "ok", "db": "connected"})
    except Exception as e:
        return JSONResponse({"status": "error", "db": str(e)}, status_code=500)
